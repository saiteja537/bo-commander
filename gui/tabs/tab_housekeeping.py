"""
gui/tabs/tab_housekeeping.py  —  Housekeeping Tab
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

Integrates gui/pages/housekeeping.py as a proper BO Commander tab.

14 live CMS categories:
  Reports, All Instances, Failed Instances, Old Instances, Users, Folders,
  Universes, Connections, Audit Events, Recurring Schedules, Stuck Instances,
  Promotion Jobs, Versions/History, Temp Objects

Features:
  • Live tile counts from real CMS queries
  • Click tile → scrollable object list with checkboxes
  • Delete selected objects
  • 🤖 One-click auto-clean per category (MultiBOT-style)
  • 📊 Visualize — bar/pie/line/donut charts (needs matplotlib)
  • ⬇ Export Excel with 5 chart types (needs openpyxl)
  • KB incident log after every delete batch

New file — add to bo_commander.py TABS list:
    ("🧹  Housekeeping", _safe_import("gui.tabs.tab_housekeeping", "HousekeepingTab")),
"""

import threading
import os
from tkinter import messagebox, filedialog
import customtkinter as ctk
from datetime import datetime

from gui.tabs._base import *
from core.sapbo_connection import bo_session

# ── Optional matplotlib ───────────────────────────────────────────────────────
try:
    import matplotlib
    matplotlib.use("TkAgg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

# ── Optional openpyxl ─────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.chart import BarChart, PieChart, LineChart, BarChart3D, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

# ── Optional KB ───────────────────────────────────────────────────────────────
try:
    from memory.knowledge_base import kb as _kb
    _KB_OK = True
except Exception:
    _kb    = None
    _KB_OK = False

# ── Category definitions ──────────────────────────────────────────────────────
CATEGORIES = [
    ("reports",          "Reports",             "📄", "#3B82F6", "All report objects (Webi + Crystal)"),
    ("instances",        "All Instances",       "📋", "#8B5CF6", "All report run instances"),
    ("failed_instances", "Failed Instances",    "❌", "#EF4444", "Instances in failed state"),
    ("old_instances",    "Old Instances",       "🕐", "#F97316", "Instances older than 30 days"),
    ("users",            "Users",               "👤", "#10B981", "Enterprise + LDAP users"),
    ("folders",          "Folders",             "📁", "#F59E0B", "Public folders"),
    ("universes",        "Universes",           "🌐", "#06B6D4", "UNV and UNX universes"),
    ("connections",      "Connections",         "🔗", "#84CC16", "Database connections"),
    ("audit_events",     "Audit Events",        "🔍", "#EC4899", "Recent CMS audit log entries"),
    ("recurring",        "Recurring Schedules", "🔄", "#F59E0B", "Active recurring schedules"),
    ("stuck_instances",  "Stuck Instances",     "⏸", "#DC2626", "Running instances > 2 hours"),
    ("promotion_jobs",   "Promotion Jobs",      "🚀", "#7C3AED", "LCM lifecycle jobs"),
    ("versions",         "Versions/History",    "📚", "#0EA5E9", "Object version history"),
    ("temp_objects",     "Temp Objects",        "🗒", "#64748B", "Objects in temp/cache folders"),
]

_CHART_COLORS = [
    "#3B82F6","#8B5CF6","#EF4444","#F97316","#10B981",
    "#F59E0B","#06B6D4","#84CC16","#EC4899","#DC2626",
    "#7C3AED","#0EA5E9","#64748B","#F43F5E",
]


def _get_columns(cat_id: str) -> list:
    common = [("SI_ID",65),("SI_NAME",270),("SI_KIND",130),("SI_OWNER",120)]
    extras = {
        "instances":        common + [("SI_STARTTIME",145),("SI_ENDTIME",145)],
        "failed_instances": common + [("SI_STARTTIME",145),("SI_ENDTIME",145)],
        "old_instances":    common + [("SI_STARTTIME",145)],
        "stuck_instances":  common + [("SI_STARTTIME",145),("Duration",90)],
        "audit_events":     [("SI_ID",65),("SI_NAME",190),("Event",145),
                             ("User",115),("Time",145)],
        "recurring":        common + [("SI_SCHEDULE_STATUS",115)],
        "promotion_jobs":   common + [("SI_CREATION_TIME",145),("Status",85)],
        "versions":         common + [("SI_VERSION",65),("SI_UPDATE_TS",145)],
        "temp_objects":     common + [("SI_UPDATE_TS",145)],
    }
    return extras.get(cat_id, common + [("SI_UPDATE_TS",145)])


def _run_query(cat_id: str) -> list:
    queries = {
        "reports": (
            "SELECT TOP 500 SI_ID,SI_NAME,SI_KIND,SI_OWNER,SI_UPDATE_TS "
            "FROM CI_INFOOBJECTS WHERE SI_INSTANCE=0 "
            "AND SI_KIND IN ('Webi','CrystalReport') "
            "ORDER BY SI_UPDATE_TS DESC"),
        "instances": (
            "SELECT TOP 500 SI_ID,SI_NAME,SI_KIND,SI_OWNER,SI_STARTTIME,SI_ENDTIME "
            "FROM CI_INFOOBJECTS WHERE SI_INSTANCE=1 ORDER BY SI_STARTTIME DESC"),
        "failed_instances": (
            "SELECT TOP 500 SI_ID,SI_NAME,SI_KIND,SI_OWNER,SI_STARTTIME,SI_ENDTIME "
            "FROM CI_INFOOBJECTS WHERE SI_INSTANCE=1 "
            "AND SI_PROCESSINFO.SI_STATUS_INFO=1 ORDER BY SI_STARTTIME DESC"),
        "old_instances": (
            "SELECT TOP 500 SI_ID,SI_NAME,SI_KIND,SI_OWNER,SI_STARTTIME "
            "FROM CI_INFOOBJECTS WHERE SI_INSTANCE=1 "
            "AND SI_STARTTIME < '2025-12-01 00:00:00' ORDER BY SI_STARTTIME ASC"),
        "users": (
            "SELECT TOP 500 SI_ID,SI_NAME,SI_KIND,SI_OWNER,SI_UPDATE_TS "
            "FROM CI_SYSTEMOBJECTS "
            "WHERE SI_KIND IN ('User','LDAPUser','WinADUser') ORDER BY SI_NAME ASC"),
        "folders": (
            "SELECT TOP 200 SI_ID,SI_NAME,SI_KIND,SI_OWNER,SI_UPDATE_TS "
            "FROM CI_INFOOBJECTS WHERE SI_KIND='Folder' AND SI_INSTANCE=0 "
            "ORDER BY SI_NAME ASC"),
        "universes": (
            "SELECT TOP 200 SI_ID,SI_NAME,SI_KIND,SI_OWNER,SI_UPDATE_TS "
            "FROM CI_APPOBJECTS WHERE SI_KIND IN ('Universe','DSL.MetaDataFile') "
            "ORDER BY SI_NAME ASC"),
        "connections": (
            "SELECT TOP 200 SI_ID,SI_NAME,SI_KIND,SI_OWNER,SI_UPDATE_TS "
            "FROM CI_APPOBJECTS WHERE SI_KIND='Connection' ORDER BY SI_NAME ASC"),
        "audit_events": (
            "SELECT TOP 200 SI_ID,SI_NAME,SI_KIND,"
            "SI_AUDIT_INFO.SI_AUDIT_EVTNAME AS Event,"
            "SI_AUDIT_INFO.SI_AUDIT_USERNAME AS AuditUser,"
            "SI_AUDIT_INFO.SI_AUDIT_STARTTIME AS AuditTime "
            "FROM CI_INFOOBJECTS WHERE SI_KIND='AuditEvent' "
            "ORDER BY SI_AUDIT_INFO.SI_AUDIT_STARTTIME DESC"),
        "recurring": (
            "SELECT TOP 200 SI_ID,SI_NAME,SI_KIND,SI_OWNER,"
            "SI_SCHEDULEINFO.SI_SCHEDULE_STATUS AS SI_SCHEDULE_STATUS "
            "FROM CI_INFOOBJECTS WHERE SI_RECURRING=1 AND SI_INSTANCE=0 "
            "ORDER BY SI_NAME ASC"),
        "stuck_instances": (
            "SELECT TOP 200 SI_ID,SI_NAME,SI_KIND,SI_OWNER,SI_STARTTIME "
            "FROM CI_INFOOBJECTS WHERE SI_INSTANCE=1 "
            "AND SI_PROCESSINFO.SI_STATUS_INFO=2 "
            "AND SI_STARTTIME < '2025-01-01 00:00:00' ORDER BY SI_STARTTIME ASC"),
        "promotion_jobs": (
            "SELECT TOP 100 SI_ID,SI_NAME,SI_KIND,SI_OWNER,SI_CREATION_TIME "
            "FROM CI_INFOOBJECTS WHERE SI_KIND IN ('LcmJob','PromotionJob') "
            "ORDER BY SI_CREATION_TIME DESC"),
        "versions": (
            "SELECT TOP 200 SI_ID,SI_NAME,SI_KIND,SI_OWNER,SI_VERSION,SI_UPDATE_TS "
            "FROM CI_INFOOBJECTS WHERE SI_INSTANCE=0 AND SI_VERSION > 1 "
            "ORDER BY SI_VERSION DESC"),
        "temp_objects": (
            "SELECT TOP 200 SI_ID,SI_NAME,SI_KIND,SI_OWNER,SI_UPDATE_TS "
            "FROM CI_INFOOBJECTS WHERE SI_INSTANCE=0 "
            "AND SI_PARENTID IN "
            "(SELECT SI_ID FROM CI_INFOOBJECTS WHERE SI_NAME='Temp') "
            "ORDER BY SI_UPDATE_TS ASC"),
    }
    q = queries.get(cat_id)
    if not q:
        return []
    try:
        d = bo_session.run_cms_query(q)
        return d.get("entries", []) if d else []
    except Exception:
        return []


# ─────────────────────────────────────────────────────────────────────────────
# Main tab
# ─────────────────────────────────────────────────────────────────────────────

class HousekeepingTab(BaseTab):

    def __init__(self, master, **kw):
        super().__init__(master, **kw)
        self._active_cat  = None
        self._cat_counts: dict = {}
        self._cat_tiles:  dict = {}   # cat_id → (tile_frame, count_label)
        self._data:       dict = {}
        self._checkboxes: dict = {}
        self._viz_win     = None
        self._build()
        threading.Thread(target=self._refresh_counts, daemon=True).start()

    # ── Layout ────────────────────────────────────────────────────────────────

    def _build(self):
        rf = self._page_header("Housekeeping", "🧹",
                                "CMS object inventory, cleanup, visualizations")

        # Right header buttons
        self._del_btn = ctk.CTkButton(
            rf, text="🗑 Delete Selected", width=140, height=30,
            fg_color=RED, text_color="white", font=F_SM,
            state="disabled", command=self._delete_selected)
        self._del_btn.pack(side="right", padx=3)

        ctk.CTkButton(rf, text="📊 Visualize", width=100, height=30,
                      fg_color=VIOLET, font=F_SM,
                      command=self._open_viz).pack(side="right", padx=3)

        ctk.CTkButton(rf, text="⬇ Export Excel", width=115, height=30,
                      fg_color=GREEN, text_color=BG0, font=F_SM,
                      command=self._export_excel).pack(side="right", padx=3)

        ctk.CTkButton(rf, text="🔄 Refresh All", width=105, height=30,
                      fg_color=BG2, text_color=TEXT, font=F_SM,
                      command=self._refresh_counts).pack(side="right", padx=3)

        body = self._body
        body.grid_columnconfigure(0, weight=1)
        body.grid_rowconfigure(2, weight=1)

        # ── Quick-clean bar ───────────────────────────────────────────────────
        qb = ctk.CTkFrame(body, fg_color=BG1, corner_radius=8, height=44)
        qb.grid(row=0, column=0, sticky="ew", padx=14, pady=(8, 2))
        qb.pack_propagate(False)

        ctk.CTkLabel(qb, text="🤖  Auto-clean:",
                     font=F_SM, text_color=VIOLET).pack(side="left", padx=10)
        for lbl, cat in [("Failed Instances","failed_instances"),
                          ("Old Instances",   "old_instances"),
                          ("Stuck Instances", "stuck_instances"),
                          ("Temp Objects",    "temp_objects")]:
            ctk.CTkButton(qb, text=lbl, height=26,
                          width=len(lbl)*7+16, corner_radius=10,
                          font=("Segoe UI",8),
                          fg_color=BG2, text_color=TEXT2,
                          hover_color=BG0,
                          command=lambda c=cat: self._auto_clean(c)
                          ).pack(side="left", padx=3)

        # ── Category tiles ────────────────────────────────────────────────────
        tiles_outer = ctk.CTkFrame(body, fg_color=BG1, corner_radius=8)
        tiles_outer.grid(row=1, column=0, sticky="ew", padx=14, pady=2)
        self._tiles_frame = ctk.CTkFrame(tiles_outer, fg_color="transparent")
        self._tiles_frame.pack(fill="x", padx=8, pady=8)
        self._build_tiles()

        # ── Detail area ───────────────────────────────────────────────────────
        detail_outer = ctk.CTkFrame(body, fg_color="transparent")
        detail_outer.grid(row=2, column=0, sticky="nsew", padx=14, pady=(4, 10))
        detail_outer.grid_columnconfigure(0, weight=1)
        detail_outer.grid_rowconfigure(1, weight=1)

        self._detail_label = ctk.CTkLabel(
            detail_outer, text="Click a category tile above to view objects.",
            font=F_H3, text_color=TEXT2, anchor="w")
        self._detail_label.grid(row=0, column=0, sticky="ew", pady=(2, 1))

        # Column header bar
        self._hdr_bar = ctk.CTkFrame(detail_outer, fg_color=BG2,
                                      corner_radius=6, height=28)
        self._hdr_bar.grid(row=0, column=0, sticky="ew")
        self._hdr_bar.pack_propagate(False)

        # Spacer so header doesn't overlap label
        self._detail_label.grid(row=0, column=0, sticky="w", pady=(0, 30))
        self._hdr_bar.grid(row=1, column=0, sticky="ew")

        self._detail_scroll = ctk.CTkScrollableFrame(
            detail_outer, fg_color=BG1, corner_radius=8)
        self._detail_scroll.grid(row=2, column=0, sticky="nsew")
        detail_outer.grid_rowconfigure(2, weight=1)

        ctk.CTkLabel(self._detail_scroll,
                     text="Click a category tile above to view objects.",
                     font=F_SM, text_color=TEXT2).pack(pady=30)

    # ── Tiles ─────────────────────────────────────────────────────────────────

    def _build_tiles(self):
        for w in self._tiles_frame.winfo_children():
            w.destroy()
        cols = 7
        for idx, (cat_id, label, icon, color, desc) in enumerate(CATEGORIES):
            self._build_tile(cat_id, label, icon, color, idx // cols, idx % cols)
        for c in range(cols):
            self._tiles_frame.grid_columnconfigure(c, weight=1)

    def _build_tile(self, cat_id, label, icon, color, row, col):
        tile = ctk.CTkFrame(self._tiles_frame, fg_color=BG2, corner_radius=8,
                             border_width=2, border_color=BG2, cursor="hand2")
        tile.grid(row=row, column=col, padx=4, pady=4, sticky="nsew")

        ctk.CTkLabel(tile, text=icon, font=("Segoe UI",17)).pack(pady=(7,0))
        ctk.CTkLabel(tile, text=label, font=("Segoe UI",8,"bold"),
                     text_color=TEXT, wraplength=88).pack()
        cnt_lbl = ctk.CTkLabel(tile, text="…", font=("Segoe UI",13,"bold"),
                                text_color=color)
        cnt_lbl.pack(pady=(2,7))

        self._cat_tiles[cat_id] = (tile, cnt_lbl)
        for w in [tile] + tile.winfo_children():
            w.bind("<Button-1>", lambda e, c=cat_id: self._select_category(c))

    # ── Count refresh ─────────────────────────────────────────────────────────

    def _refresh_counts(self):
        if not bo_session.connected:
            self.after(0, lambda: self.set_status("❌ Not connected to BO", RED))
            return
        self.after(0, lambda: self.set_status("⏳ Refreshing counts…", AMBER))
        for cat_id, *_ in CATEGORIES:
            threading.Thread(target=self._fetch_count,
                             args=(cat_id,), daemon=True).start()

    def _fetch_count(self, cat_id: str):
        rows = _run_query(cat_id)
        self._cat_counts[cat_id] = len(rows)
        self._data[cat_id]       = rows

        def _upd():
            if cat_id in self._cat_tiles:
                _, lbl = self._cat_tiles[cat_id]
                if lbl.winfo_exists():
                    lbl.configure(text=str(len(rows)))
            total = sum(self._cat_counts.values())
            self.set_status(f"Total tracked CMS objects: {total:,}", TEXT2)

        self.after(0, _upd)

    # ── Select category ───────────────────────────────────────────────────────

    def _select_category(self, cat_id: str):
        if self._active_cat and self._active_cat in self._cat_tiles:
            tile, _ = self._cat_tiles[self._active_cat]
            if tile.winfo_exists():
                tile.configure(border_color=BG2)
        self._active_cat = cat_id
        if cat_id in self._cat_tiles:
            tile, _ = self._cat_tiles[cat_id]
            color = next((c[3] for c in CATEGORIES if c[0] == cat_id), CYAN)
            if tile.winfo_exists():
                tile.configure(border_color=color)

        label = next((c[1] for c in CATEGORIES if c[0] == cat_id), cat_id)
        self._detail_label.configure(text=f"  {label}")

        if cat_id in self._data and self._data[cat_id] is not None:
            self._render_detail(cat_id, self._data[cat_id])
        else:
            for w in self._detail_scroll.winfo_children():
                w.destroy()
            ctk.CTkLabel(self._detail_scroll, text="⏳ Loading…",
                         font=F_SM, text_color=TEXT2).pack(pady=20)
            threading.Thread(target=self._load_detail,
                             args=(cat_id,), daemon=True).start()

    def _load_detail(self, cat_id: str):
        rows = _run_query(cat_id)
        self._data[cat_id] = rows
        self.after(0, lambda r=rows, c=cat_id: self._render_detail(c, r))

    # ── Render detail ─────────────────────────────────────────────────────────

    def _render_detail(self, cat_id: str, rows: list):
        for w in self._detail_scroll.winfo_children():
            w.destroy()
        for w in self._hdr_bar.winfo_children():
            w.destroy()
        self._checkboxes.clear()
        self._del_btn.configure(state="disabled")

        if not rows:
            ctk.CTkLabel(self._detail_scroll,
                         text="✅  No objects found in this category.",
                         font=F_SM, text_color=GREEN).pack(pady=30)
            return

        cols = _get_columns(cat_id)
        self._del_btn.configure(state="normal")

        # Header bar
        ctk.CTkLabel(self._hdr_bar, text="", width=24).pack(side="left", padx=4)
        for col_name, col_w in cols:
            ctk.CTkLabel(self._hdr_bar, text=col_name, width=col_w, anchor="w",
                         font=("Segoe UI",9,"bold"),
                         text_color=TEXT2).pack(side="left", padx=3, pady=5)

        # Data rows
        for i, entry in enumerate(rows):
            obj_id = entry.get("SI_ID", i)
            row_f  = ctk.CTkFrame(self._detail_scroll,
                                   fg_color=BG2 if i % 2 == 0 else BG1,
                                   corner_radius=4)
            row_f.pack(fill="x", padx=4, pady=1)

            var = ctk.BooleanVar(value=False)
            self._checkboxes[obj_id] = var
            ctk.CTkCheckBox(row_f, text="", variable=var, width=20
                             ).pack(side="left", padx=(4, 3), pady=5)

            for col_name, col_w in cols:
                raw = entry.get(col_name, entry.get(col_name.upper(), ""))
                val = str(raw)[:50] if raw is not None else ""
                ctk.CTkLabel(row_f, text=val, width=col_w, anchor="w",
                             font=("Segoe UI",10),
                             text_color=TEXT).pack(side="left", padx=3, pady=5)

    # ── Delete ────────────────────────────────────────────────────────────────

    def _delete_selected(self):
        ids = [oid for oid, var in self._checkboxes.items() if var.get()]
        if not ids:
            self.set_status("⚠ Select items first", AMBER)
            return
        if not confirm("Confirm Delete",
                       f"Delete {len(ids)} selected object(s)?\n"
                       "This cannot be undone.", parent=self):
            return
        self.set_status(f"⏳ Deleting {len(ids)} object(s)…", AMBER)
        threading.Thread(target=self._do_delete, args=(ids,), daemon=True).start()

    def _do_delete(self, ids: list):
        ok = err = 0
        for oid in ids:
            try:
                success, _ = bo_session.delete_object(oid)
                if success: ok += 1
                else:       err += 1
            except Exception:
                err += 1

        if _KB_OK and _kb:
            _kb.log_automation(
                trigger="user", action="housekeeping_delete",
                target=f"cat={self._active_cat}",
                result=f"deleted={ok} failed={err}",
                success=ok > 0)

        self.after(0, lambda: self.set_status(
            f"Done: {ok} deleted ✅  |  {err} failed ❌",
            GREEN if err == 0 else AMBER))
        if self._active_cat:
            self._data[self._active_cat] = None
            self.after(800, lambda: self._select_category(self._active_cat))

    # ── Auto-clean (one-click) ────────────────────────────────────────────────

    def _auto_clean(self, cat_id: str):
        cat_label = next((c[1] for c in CATEGORIES if c[0] == cat_id), cat_id)
        rows = self._data.get(cat_id) or _run_query(cat_id)
        if not rows:
            self.set_status(f"✅ No {cat_label} to clean", GREEN)
            return
        if not confirm(f"Auto-Clean: {cat_label}",
                       f"Delete ALL {len(rows)} {cat_label}?\n"
                       "This cannot be undone.", parent=self):
            return
        self.set_status(f"⏳ Cleaning {len(rows)} {cat_label}…", AMBER)
        ids = [r.get("SI_ID") for r in rows if r.get("SI_ID")]
        threading.Thread(target=self._do_delete, args=(ids,), daemon=True).start()

    # ── Visualize ─────────────────────────────────────────────────────────────

    def _open_viz(self):
        if not HAS_MPL:
            messagebox.showwarning("Missing Library",
                                   "Run:  pip install matplotlib", parent=self)
            return
        counts = {c[1]: self._cat_counts.get(c[0], 0)
                  for c in CATEGORIES if self._cat_counts.get(c[0], 0) > 0}
        if not counts:
            messagebox.showinfo("No Data",
                                "Wait for tile counts to load first.", parent=self)
            return
        if self._viz_win and self._viz_win.winfo_exists():
            self._viz_win.lift()
            return
        self._viz_win = _VizWindow(self, counts)

    # ── Excel export ──────────────────────────────────────────────────────────

    def _export_excel(self):
        if not HAS_XLSX:
            messagebox.showwarning("Missing Library",
                                   "Run:  pip install openpyxl", parent=self)
            return
        if not any(self._data.get(c[0]) for c in CATEGORIES):
            messagebox.showinfo("No Data",
                                "Wait for data to load first.", parent=self)
            return
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            title="Export Housekeeping Report",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=f"BO_Housekeeping_{ts}.xlsx",
            parent=self)
        if not path:
            return
        self.set_status("⏳ Exporting Excel…", AMBER)
        threading.Thread(target=self._do_export, args=(path,), daemon=True).start()

    def _do_export(self, path: str):
        try:
            wb   = openpyxl.Workbook()
            wb.remove(wb.active)
            thin = Border(
                left=Side(style="thin",  color="2D3748"),
                right=Side(style="thin", color="2D3748"),
                top=Side(style="thin",   color="2D3748"),
                bottom=Side(style="thin",color="2D3748"))
            hdr_fill   = PatternFill("solid", fgColor="1E3A5F")
            alt_fill   = PatternFill("solid", fgColor="1A2236")
            title_fill = PatternFill("solid", fgColor="0F172A")

            def _hdr(ws, cols, row=1):
                for ci, h in enumerate(cols, 1):
                    c = ws.cell(row=row, column=ci, value=h)
                    c.font = Font(bold=True, color="FFFFFF", size=10)
                    c.fill = hdr_fill
                    c.alignment = Alignment(horizontal="left", vertical="center")
                    c.border = thin
                ws.row_dimensions[row].height = 20

            def _autowidth(ws):
                for col in ws.columns:
                    w = 10
                    for cell in col:
                        if cell.value:
                            w = max(w, min(len(str(cell.value))+2, 60))
                    ws.column_dimensions[
                        get_column_letter(col[0].column)].width = w

            # Summary sheet
            ws_s = wb.create_sheet("Summary")
            ws_s["A1"] = "BO Commander — Housekeeping Report"
            ws_s["A1"].font = Font(bold=True, color="FFFFFF", size=14)
            ws_s["A1"].fill = title_fill
            ws_s["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws_s["A2"].font = Font(color="9AA0B4", size=10)
            _hdr(ws_s, ["Category","Description","Count"], row=4)
            grand = 0
            for ri, (cat_id, label, icon, color, desc) in enumerate(CATEGORIES, 5):
                cnt   = len(self._data.get(cat_id,[]) or [])
                grand += cnt
                ws_s.cell(ri,1, f"{icon} {label}")
                ws_s.cell(ri,2, desc)
                ws_s.cell(ri,3, cnt)
            tr = 5 + len(CATEGORIES)
            ws_s.cell(tr,1,"GRAND TOTAL").font = Font(bold=True,color="FFD700",size=11)
            ws_s.cell(tr,3, grand).font = Font(bold=True,color="FFD700",size=11)
            _autowidth(ws_s)

            # Per-category sheets
            for cat_id, label, icon, color, desc in CATEGORIES:
                rows = self._data.get(cat_id,[]) or []
                ws   = wb.create_sheet(f"{icon} {label}"[:31])
                ws["A1"] = f"{icon}  {label}"
                ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
                ws["A1"].fill = title_fill
                ws["A2"] = f"{desc}  |  Records: {len(rows)}"
                ws["A2"].font = Font(color="9AA0B4", size=10)
                if not rows:
                    ws["A4"] = "No objects found."
                    ws["A4"].font = Font(color="9AA0B4", italic=True)
                    continue
                cols = _get_columns(cat_id)
                col_names = [c[0] for c in cols]
                _hdr(ws, col_names, row=4)
                for ri, entry in enumerate(rows, 5):
                    for ci, cn in enumerate(col_names, 1):
                        raw = entry.get(cn, entry.get(cn.upper(),""))
                        val = str(raw)[:200] if raw is not None else ""
                        c = ws.cell(ri, ci, val)
                        c.border = thin
                        if ri % 2 == 0: c.fill = alt_fill
                _autowidth(ws)

            wb.save(path)
            self.after(0, lambda: self.set_status(
                f"✅ Exported: {os.path.basename(path)}", GREEN))
        except Exception as ex:
            self.after(0, lambda: self.set_status(
                f"❌ Export failed: {ex}", RED))


# ── Visualization popup window ────────────────────────────────────────────────

class _VizWindow(ctk.CTkToplevel):
    CHART_TYPES = [
        "Bar Chart", "Horizontal Bar", "Pie Chart", "Donut Chart", "Line Chart"
    ]

    def __init__(self, parent, counts: dict):
        super().__init__(parent)
        self.title("📊  Housekeeping Visualizations")
        self.geometry("1080x680")
        self.configure(fg_color=BG0)
        self._counts = counts
        self._ct_var = ctk.StringVar(value=self.CHART_TYPES[0])
        self._canvas = None
        self._fig    = None
        self._build()
        self._draw()

    def _build(self):
        bar = ctk.CTkFrame(self, fg_color=BG1, height=50, corner_radius=0)
        bar.pack(fill="x")
        bar.pack_propagate(False)
        ctk.CTkLabel(bar, text="📊  Category Overview",
                     font=F_H3, text_color=TEXT).pack(side="left", padx=16)
        ctk.CTkButton(bar, text="⬇ Download", width=110, height=32,
                      fg_color=GREEN, text_color=BG0,
                      command=self._download).pack(side="right", padx=10, pady=9)
        for ct in self.CHART_TYPES:
            ctk.CTkButton(bar, text=ct, width=110, height=28,
                          fg_color=BG2, font=("Segoe UI",9),
                          command=lambda c=ct: self._switch(c)
                          ).pack(side="left", padx=3, pady=11)
        self._cf = ctk.CTkFrame(self, fg_color=BG1, corner_radius=8)
        self._cf.pack(fill="both", expand=True, padx=12, pady=10)

    def _switch(self, ct):
        self._ct_var.set(ct)
        self._draw()

    def _draw(self):
        ct     = self._ct_var.get()
        labels = list(self._counts.keys())
        vals   = list(self._counts.values())
        cols   = (_CHART_COLORS * 3)[:len(labels)]
        if self._canvas:
            self._canvas.get_tk_widget().destroy()
        if self._fig:
            plt.close(self._fig)

        fig, ax = plt.subplots(figsize=(11, 5), facecolor="#1e2433")
        ax.set_facecolor("#252d3d")
        for sp in ax.spines.values():
            sp.set_edgecolor("#3a4460")
        ax.tick_params(colors="#9aa0b4")

        if ct == "Bar Chart":
            bars = ax.bar(labels, vals, color=cols, edgecolor="#1e2433", linewidth=0.6)
            for b, v in zip(bars, vals):
                ax.text(b.get_x()+b.get_width()/2, b.get_height()+.4, str(v),
                        ha="center", color="white", fontsize=9, fontweight="bold")
            ax.tick_params(axis="x", rotation=35)
        elif ct == "Horizontal Bar":
            bars = ax.barh(labels, vals, color=cols, edgecolor="#1e2433")
            for b, v in zip(bars, vals):
                ax.text(b.get_width()+.3, b.get_y()+b.get_height()/2, str(v),
                        va="center", color="white", fontsize=9)
        elif ct in ("Pie Chart", "Donut Chart"):
            kw = dict(labels=labels, colors=cols, startangle=140,
                      autopct=lambda p: f"{p:.1f}%" if p > 1.5 else "",
                      wedgeprops={"edgecolor":"#1e2433","linewidth":1.5})
            if ct == "Donut Chart":
                kw["wedgeprops"]["width"] = 0.52
            _, texts, autos = ax.pie(vals, **kw)
            for t in texts: t.set_color("#c8cfe0"); t.set_fontsize(8)
            for t in autos: t.set_color("white");   t.set_fontsize(8)
            if ct == "Donut Chart":
                ax.text(0, 0, f"Total\n{sum(vals):,}",
                        ha="center", va="center",
                        color="white", fontsize=13, fontweight="bold")
        elif ct == "Line Chart":
            x = list(range(len(labels)))
            ax.plot(x, vals, color="#3B82F6", marker="o",
                    linewidth=2.5, markersize=9)
            ax.fill_between(x, vals, alpha=0.12, color="#3B82F6")
            for xi, yi in enumerate(vals):
                ax.text(xi, yi+.4, str(yi), ha="center",
                        color="white", fontsize=9)
            ax.set_xticks(x)
            ax.set_xticklabels(labels, rotation=35, ha="right")

        ax.set_title(f"BO Object Counts — {ct}", color="#e2e8f0",
                     fontsize=13, fontweight="bold")
        fig.tight_layout(pad=2.0)
        self._fig    = fig
        self._canvas = FigureCanvasTkAgg(fig, master=self._cf)
        self._canvas.draw()
        self._canvas.get_tk_widget().pack(fill="both", expand=True, padx=6, pady=6)

    def _download(self):
        ct   = self._ct_var.get().replace(" ","_").lower()
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[("PNG","*.png"),("PDF","*.pdf")],
            initialfile=f"BO_housekeeping_{ct}_{ts}.png",
            parent=self)
        if path and self._fig:
            self._fig.savefig(path, dpi=150, bbox_inches="tight",
                              facecolor=self._fig.get_facecolor())

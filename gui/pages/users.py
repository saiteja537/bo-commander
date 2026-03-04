"""
users.py  —  Enterprise User & Group Command Center
Features beyond CMC:
  • Summary insight cards (Total, Active, Disabled, Locked, External, Never Logged In)
  • Live search + filter by Status / Auth Type
  • Risk indicator badges (Never Logged In, Inactive, Locked, External)
  • Multi-select + bulk actions (Enable, Disable, Reset Password, Export Selected)
  • Full detail drawer (500 px, no truncation) with tabs:
      Overview | Groups & Roles | Security | Activity
  • Group tab: list + member detail
  • Hierarchy tab: group membership tree
  • Export: Excel (full data + charts), CSV, JSON
  • Import wizard: Upload → Map → Validate → Preview → Apply
"""

import threading
import csv
import json
import io
import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime

from config import Config
from core.sapbo_connection import bo_session

C = Config.COLORS
F = Config.FONTS

_PAGE_REF = [None]

AUTH_MAP = {
    'secEnterprise': 'Enterprise',
    'secLDAP':       'LDAP',
    'secWinAD':      'Windows AD',
    'secWindows':    'Windows AD',
    'secSAPR3':      'SAP',
}
STATUS_COLOR  = {'Enabled': C['success'], 'Disabled': C['danger'], 'Locked': C['warning']}
RISK_COLORS   = {'Never Logged In': '#F59E0B', 'Inactive': '#F97316',
                 'Locked': '#EF4444',          'External': '#8B5CF6'}


def _fmt(val):
    if not val:
        return '—'
    s = str(val)
    try:
        return datetime.strptime(s[:19].replace('T', ' '),
                                 '%Y-%m-%d %H:%M:%S').strftime('%d %b %Y  %H:%M')
    except Exception:
        return s[:16] or '—'


def _risk_tags(u):
    tags = []
    if not u.get('date_modified') and not u.get('last_login'):
        tags.append('Never Logged In')
    if u.get('account_status') == 'Locked':
        tags.append('Locked')
    if u.get('auth_type', 'Enterprise') not in ('Enterprise', ''):
        tags.append('External')
    return tags


def _run_bg(fn, cb):
    root = _PAGE_REF[0]
    def _w():
        try: r = fn()
        except Exception: r = None
        if root:
            try:
                def _safe(res=r):
                    try: cb(res)
                    except Exception: pass
                root.after(0, _safe)
            except Exception: pass
    threading.Thread(target=_w, daemon=True).start()


# ─── Treeview factory ─────────────────────────────────────────────────────────

def _tree(parent, cols, row_h=32):
    """cols = [(id, heading, width, stretch), ...]"""
    sn = f'U{id(parent)}.TV'
    s = ttk.Style()
    s.configure(sn,
                background=C['bg_secondary'],
                foreground=C['text_primary'],
                fieldbackground=C['bg_secondary'],
                rowheight=row_h,
                font=('Segoe UI', 11),
                borderwidth=0)
    s.configure(f'{sn}.Heading',
                background=C['bg_tertiary'],
                foreground=C['text_secondary'],
                font=('Segoe UI', 10, 'bold'),
                relief='flat', borderwidth=0)
    s.map(sn,
          background=[('selected', C['primary'])],
          foreground=[('selected', 'white')])
    s.layout(sn, [('Treeview.treearea', {'sticky': 'nswe'})])

    tv = ttk.Treeview(parent, style=sn, show='headings', selectmode='extended',
                      columns=[c[0] for c in cols])
    for cid, hdr, w, stretch in cols:
        tv.heading(cid, text=hdr)
        tv.column(cid, width=w, minwidth=40, stretch=stretch)

    vsb = ctk.CTkScrollbar(parent, orientation='vertical',   command=tv.yview)
    hsb = ctk.CTkScrollbar(parent, orientation='horizontal',  command=tv.xview)
    tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    hsb.pack(side='bottom', fill='x')
    vsb.pack(side='right',  fill='y')
    tv.pack(side='left', fill='both', expand=True)
    return tv


# ─── Insight card ─────────────────────────────────────────────────────────────

def _card(parent, label, value='0', color=None):
    color = color or C['text_primary']
    fr = ctk.CTkFrame(parent, fg_color=C['bg_secondary'],
                      corner_radius=10, width=130, height=70)
    fr.pack(side='left', padx=5, pady=4)
    fr.pack_propagate(False)
    ctk.CTkLabel(fr, text=label, font=('Segoe UI', 9),
                 text_color=C['text_secondary']).pack(pady=(10, 0))
    lbl = ctk.CTkLabel(fr, text=str(value), font=('Segoe UI', 22, 'bold'),
                       text_color=color)
    lbl.pack()
    return lbl


# ─── Detail Drawer ────────────────────────────────────────────────────────────

class _UserDrawer(ctk.CTkToplevel):
    """Floating detail window — no layout constraints, values never clip."""

    TABS = ['Overview', 'Groups & Roles', 'Security', 'Activity']

    def __init__(self, master):
        super().__init__(master)
        self.title('User Details')
        self.geometry('560x720')
        self.resizable(True, True)
        self.configure(fg_color=C['bg_primary'])
        self.attributes('-topmost', False)
        self._uid = None
        self._user = {}
        self._build()

    def _build(self):
        # ── Title bar ─────────────────────────────────────────────────────────
        tb = ctk.CTkFrame(self, fg_color=C['bg_tertiary'], corner_radius=0, height=48)
        tb.pack(fill='x')
        tb.pack_propagate(False)
        self._title = ctk.CTkLabel(tb, text='', font=('Segoe UI', 14, 'bold'),
                                   text_color=C['text_primary'], anchor='w')
        self._title.pack(side='left', padx=16, fill='x', expand=True)
        self._status_badge = ctk.CTkLabel(tb, text='', font=('Segoe UI', 11),
                                          width=80, corner_radius=6)
        self._status_badge.pack(side='right', padx=12)
        ctk.CTkButton(tb, text='✕', width=32, height=28,
                      fg_color='transparent', hover_color=C['danger'],
                      text_color=C['text_secondary'],
                      command=self.withdraw).pack(side='right', padx=4)

        # ── Tab bar ────────────────────────────────────────────────────────────
        tab_bar = ctk.CTkFrame(self, fg_color=C['bg_secondary'],
                               corner_radius=0, height=38)
        tab_bar.pack(fill='x')
        tab_bar.pack_propagate(False)
        self._tab_btns   = {}
        self._tab_frames = {}
        self._active_tab = ctk.StringVar(value='Overview')
        for name in self.TABS:
            b = ctk.CTkButton(tab_bar, text=name, width=130, height=38,
                              font=('Segoe UI', 11),
                              fg_color='transparent',
                              hover_color=C['bg_tertiary'],
                              text_color=C['text_secondary'],
                              corner_radius=0,
                              command=lambda n=name: self._switch(n))
            b.pack(side='left')
            self._tab_btns[name] = b
            self._tab_frames[name] = ctk.CTkScrollableFrame(
                self, fg_color=C['bg_primary'], corner_radius=0)

        self._switch('Overview')

    def _switch(self, name):
        self._active_tab.set(name)
        for n, f in self._tab_frames.items():
            f.pack_forget()
        self._tab_frames[name].pack(fill='both', expand=True)
        for n, b in self._tab_btns.items():
            b.configure(fg_color=C['primary'] if n == name else 'transparent',
                        text_color='white' if n == name else C['text_secondary'])
        if self._uid:
            self._load_tab(name)

    # ── Load ──────────────────────────────────────────────────────────────────

    def load(self, user):
        self._uid  = user['id']
        self._user = user
        # Header
        display = user.get('name', '?')
        full = user.get('full_name', '')
        self._title.configure(text=f'👤  {display}' + (f'  ({full})' if full else ''))
        status = user.get('account_status', 'Enabled')
        self._status_badge.configure(
            text=f'● {status}',
            text_color=STATUS_COLOR.get(status, C['text_secondary']))
        self.deiconify()
        self.lift()
        self._switch(self._active_tab.get())

    def _load_tab(self, name):
        self._clear(name)
        uid = self._uid
        if name == 'Overview':
            self._render_overview()
        elif name == 'Groups & Roles':
            _run_bg(lambda: bo_session.get_user_member_of(uid), self._render_groups)
        elif name == 'Security':
            _run_bg(lambda: bo_session.get_user_properties(uid), self._render_security)
        elif name == 'Activity':
            _run_bg(lambda: bo_session.get_user_properties(uid), self._render_activity)

    # ── Tab renders ───────────────────────────────────────────────────────────

    def _render_overview(self):
        f = self._tab('Overview')
        u = self._user

        # Risk indicators
        risks = _risk_tags(u)
        if risks:
            rb = ctk.CTkFrame(f, fg_color='transparent')
            rb.pack(fill='x', padx=16, pady=(12, 4))
            ctk.CTkLabel(rb, text='⚠  Risk Indicators:',
                         font=('Segoe UI', 10, 'bold'),
                         text_color=C['warning']).pack(side='left', padx=(0, 8))
            for r in risks:
                ctk.CTkLabel(rb, text=r,
                             font=('Segoe UI', 10),
                             fg_color=RISK_COLORS.get(r, C['bg_tertiary']),
                             text_color='white',
                             corner_radius=8,
                             padx=8, pady=2).pack(side='left', padx=3)

        self._section(f, 'Account Information')
        self._row(f, 'Account Name',  u.get('name', ''))
        self._row(f, 'Full Name',     u.get('full_name', ''))
        self._row(f, 'Email Address', u.get('email', ''))
        self._row(f, 'Description',   u.get('description', ''))

        self._section(f, 'Account Status')
        status = u.get('account_status', 'Enabled')
        row = ctk.CTkFrame(f, fg_color='transparent')
        row.pack(fill='x', padx=16, pady=4)
        ctk.CTkLabel(row, text='Account Status', width=160,
                     anchor='w', font=('Segoe UI', 11),
                     text_color=C['text_secondary']).pack(side='left')
        ctk.CTkLabel(row, text=f'● {status}',
                     font=('Segoe UI', 11, 'bold'),
                     text_color=STATUS_COLOR.get(status, C['text_secondary'])
                     ).pack(side='left')

        self._section(f, 'Authentication & Access')
        self._row(f, 'Authentication Type', u.get('auth_type', ''))
        self._row(f, 'Tenant',              u.get('tenant', ''))

        self._section(f, 'Timeline')
        self._row(f, 'Date Created',  _fmt(u.get('date_created', '')))
        self._row(f, 'Date Modified', _fmt(u.get('date_modified', '')))

        # Quick actions inside drawer
        self._section(f, 'Quick Actions')
        aq = ctk.CTkFrame(f, fg_color='transparent')
        aq.pack(fill='x', padx=16, pady=6)
        _ab = dict(height=30, corner_radius=6, font=('Segoe UI', 11), width=130)
        uid = self._uid
        ctk.CTkButton(aq, text='Disable User',
                      fg_color=C['danger'], hover_color='#DC2626',
                      command=lambda: self._quick_disable(uid), **_ab
                      ).pack(side='left', padx=(0, 6))
        ctk.CTkButton(aq, text='Reset Password',
                      fg_color=C['warning'], hover_color='#D97706',
                      command=lambda: self._quick_reset(uid), **_ab
                      ).pack(side='left')

    def _render_groups(self, groups):
        f = self._tab('Groups & Roles')
        if not groups:
            self._empty(f, 'Not a member of any group.'); return
        self._section(f, f'Member of {len(groups)} Group(s)')
        for g in groups:
            card = ctk.CTkFrame(f, fg_color=C['bg_secondary'], corner_radius=8)
            card.pack(fill='x', padx=16, pady=3)
            row = ctk.CTkFrame(card, fg_color='transparent', height=40)
            row.pack(fill='x', padx=12)
            row.pack_propagate(False)
            ctk.CTkLabel(row, text='👥', font=('Segoe UI', 14),
                         width=28).pack(side='left')
            ctk.CTkLabel(row, text=g.get('name', '?'),
                         font=('Segoe UI', 12, 'bold'),
                         text_color=C['primary'],
                         anchor='w').pack(side='left', fill='x', expand=True)
            desc = g.get('description', '')
            if desc:
                ctk.CTkLabel(card, text=desc, font=('Segoe UI', 10),
                             text_color=C['text_secondary'],
                             anchor='w',
                             wraplength=460
                             ).pack(fill='x', padx=12, pady=(0, 6))

    def _render_security(self, data):
        f = self._tab('Security')
        if not data:
            self._empty(f, 'No security data available.'); return
        self._section(f, 'Account Security Settings')
        disabled = bool(data.get('SI_DISABLED', False))
        locked   = bool(data.get('SI_PASSWORD_LOCKED', False))
        status = 'Disabled' if disabled else ('Locked' if locked else 'Enabled')
        self._row(f, 'Account Status',   status)
        self._row(f, 'Account Locked',   '✔ Yes' if locked   else '✗ No')
        self._row(f, 'Account Disabled', '✔ Yes' if disabled else '✗ No')

        auth_raw = data.get('SI_AUTH_TYPE', 'secEnterprise')
        self._section(f, 'Authentication')
        self._row(f, 'Auth Type (raw)',   auth_raw)
        self._row(f, 'Auth Type (label)', AUTH_MAP.get(auth_raw, auth_raw))

        self._section(f, 'Notes')
        ctk.CTkLabel(f,
                     text='Full security rights are managed in CMC.\n'
                          'CMC → Users and Groups → [User] → User Security.',
                     font=('Segoe UI', 10),
                     text_color=C['text_secondary'],
                     justify='left',
                     wraplength=480).pack(padx=16, pady=8, anchor='w')

    def _render_activity(self, data):
        f = self._tab('Activity')
        u = self._user
        self._section(f, 'Login Activity')
        self._row(f, 'Last Login',        _fmt(data.get('SI_LAST_LOGIN_TIME', '') if data else ''))
        self._row(f, 'Account Created',   _fmt(u.get('date_created', '')))
        self._row(f, 'Last Modified',     _fmt(u.get('date_modified', '')))

        risks = _risk_tags(u)
        self._section(f, 'Risk Assessment')
        if risks:
            for r in risks:
                row = ctk.CTkFrame(f, fg_color=C['bg_secondary'], corner_radius=6, height=34)
                row.pack(fill='x', padx=16, pady=2)
                row.pack_propagate(False)
                ctk.CTkLabel(row, text=f'⚠  {r}',
                             font=('Segoe UI', 11),
                             text_color=RISK_COLORS.get(r, C['warning'])
                             ).pack(side='left', padx=12)
        else:
            self._empty(f, '✅  No risk indicators detected.')

        self._section(f, 'Note')
        ctk.CTkLabel(f,
                     text='Detailed activity logs require SAP BO Auditing DB.\n'
                          'Enable Audit DB in Settings to see login history,\n'
                          'report access frequency, and session counts.',
                     font=('Segoe UI', 10),
                     text_color=C['text_secondary'],
                     justify='left',
                     wraplength=480).pack(padx=16, pady=8, anchor='w')

    # ── Quick actions ─────────────────────────────────────────────────────────

    def _quick_disable(self, uid):
        if messagebox.askyesno('Confirm', 'Disable this user?', parent=self):
            _run_bg(lambda: bo_session.disable_user(uid, True),
                    lambda r: messagebox.showinfo('Done', 'User disabled.', parent=self))

    def _quick_reset(self, uid):
        dlg = ctk.CTkToplevel(self)
        dlg.title('Reset Password')
        dlg.geometry('340x160')
        dlg.configure(fg_color=C['bg_primary'])
        dlg.attributes('-topmost', True)
        ctk.CTkLabel(dlg, text='New Password:', font=F['body'],
                     text_color=C['text_primary']).pack(pady=(20, 4))
        entry = ctk.CTkEntry(dlg, show='*', width=250)
        entry.pack()
        def _do():
            pwd = entry.get().strip()
            if not pwd:
                return
            _run_bg(lambda: bo_session.reset_user_password(uid, pwd),
                    lambda r: (messagebox.showinfo('Done', 'Password reset.', parent=self),
                               dlg.destroy()))
        ctk.CTkButton(dlg, text='Reset', fg_color=C['primary'],
                      command=_do, width=120).pack(pady=12)

    # ── Render helpers ────────────────────────────────────────────────────────

    def _tab(self, n):    return self._tab_frames[n]
    def _clear(self, n):
        for w in self._tab_frames[n].winfo_children():
            w.destroy()

    def _section(self, parent, text):
        ctk.CTkLabel(parent, text=text,
                     font=('Segoe UI', 11, 'bold'),
                     text_color=C['primary'],
                     anchor='w').pack(fill='x', padx=16, pady=(14, 2))
        ctk.CTkFrame(parent, height=1,
                     fg_color=C['bg_secondary']).pack(fill='x', padx=16, pady=(0, 4))

    def _row(self, parent, label, value):
        """Two-column row. Full value displayed — no clipping."""
        row = ctk.CTkFrame(parent, fg_color='transparent')
        row.pack(fill='x', padx=16, pady=3)
        ctk.CTkLabel(row, text=label, width=160, anchor='nw',
                     font=('Segoe UI', 11),
                     text_color=C['text_secondary']).pack(side='left', padx=(0, 8))
        ctk.CTkLabel(row, text=str(value) if value else '—',
                     anchor='nw', justify='left',
                     font=('Segoe UI', 11),
                     text_color=C['text_primary'],
                     wraplength=320).pack(side='left', fill='x', expand=True)

    def _empty(self, parent, text):
        ctk.CTkLabel(parent, text=text, font=('Segoe UI', 11),
                     text_color=C['text_secondary']).pack(pady=30, padx=16)


# ─── Import Wizard ────────────────────────────────────────────────────────────

class _ImportWizard(ctk.CTkToplevel):
    STEPS = ['1. Upload', '2. Map Fields', '3. Validate', '4. Preview', '5. Apply']

    def __init__(self, master, on_done=None):
        super().__init__(master)
        self.title('Import Users — Wizard')
        self.geometry('680x560')
        self.configure(fg_color=C['bg_primary'])
        self.attributes('-topmost', True)
        self._step = 0
        self._file_data  = []
        self._headers    = []
        self._mapping    = {}
        self._validated  = []
        self._on_done    = on_done
        self._build()
        self._show_step(0)

    def _build(self):
        # Progress bar
        prog = ctk.CTkFrame(self, fg_color=C['bg_secondary'], corner_radius=0, height=44)
        prog.pack(fill='x')
        prog.pack_propagate(False)
        self._step_btns = []
        for i, s in enumerate(self.STEPS):
            b = ctk.CTkButton(prog, text=s, width=120, height=44,
                              fg_color='transparent',
                              text_color=C['text_secondary'],
                              corner_radius=0, state='disabled',
                              font=('Segoe UI', 10))
            b.pack(side='left')
            self._step_btns.append(b)

        # Content area
        self._content = ctk.CTkScrollableFrame(self, fg_color=C['bg_primary'],
                                               corner_radius=0)
        self._content.pack(fill='both', expand=True, padx=20, pady=10)

        # Nav buttons
        nav = ctk.CTkFrame(self, fg_color='transparent', height=44)
        nav.pack(fill='x', padx=20, pady=(0, 12))
        nav.pack_propagate(False)
        self._back_btn = ctk.CTkButton(nav, text='← Back', width=100,
                                       fg_color=C['bg_tertiary'],
                                       hover_color=C['bg_secondary'],
                                       command=self._back, state='disabled')
        self._back_btn.pack(side='left')
        self._next_btn = ctk.CTkButton(nav, text='Next →', width=100,
                                       fg_color=C['primary'],
                                       hover_color=C['primary_hover'],
                                       command=self._next)
        self._next_btn.pack(side='right')

    def _show_step(self, n):
        self._step = n
        for w in self._content.winfo_children():
            w.destroy()
        for i, b in enumerate(self._step_btns):
            b.configure(
                fg_color=C['primary'] if i == n else 'transparent',
                text_color='white' if i == n else C['text_secondary'])
        self._back_btn.configure(state='normal' if n > 0 else 'disabled')
        self._next_btn.configure(text='Apply & Close' if n == 4 else 'Next →')
        getattr(self, f'_step{n}')()

    def _back(self): self._show_step(max(0, self._step - 1))
    def _next(self):
        if self._step == 4:
            self._apply()
        else:
            self._show_step(min(4, self._step + 1))

    # ── Steps ─────────────────────────────────────────────────────────────────

    def _step0(self):
        c = self._content
        ctk.CTkLabel(c, text='Upload a file to import users.',
                     font=F['body'], text_color=C['text_primary']).pack(pady=(20, 4))
        ctk.CTkLabel(c, text='Supported formats: Excel (.xlsx), CSV (.csv), JSON (.json)',
                     font=F['small'], text_color=C['text_secondary']).pack(pady=(0, 16))

        def _pick():
            path = filedialog.askopenfilename(
                filetypes=[('Supported', '*.xlsx *.csv *.json'),
                           ('Excel', '*.xlsx'), ('CSV', '*.csv'), ('JSON', '*.json')])
            if not path:
                return
            try:
                if path.endswith('.csv'):
                    with open(path, newline='', encoding='utf-8-sig') as fp:
                        reader = csv.DictReader(fp)
                        self._headers   = reader.fieldnames or []
                        self._file_data = list(reader)
                elif path.endswith('.json'):
                    with open(path, encoding='utf-8') as fp:
                        data = json.load(fp)
                    if isinstance(data, list) and data:
                        self._headers   = list(data[0].keys())
                        self._file_data = data
                else:
                    import openpyxl
                    wb = openpyxl.load_workbook(path, read_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if rows:
                        self._headers   = [str(h) for h in rows[0]]
                        self._file_data = [dict(zip(self._headers, r))
                                           for r in rows[1:] if any(r)]
                    wb.close()
                lbl.configure(text=f'✅  Loaded {len(self._file_data)} rows, '
                                   f'{len(self._headers)} columns:\n'
                                   f'{", ".join(self._headers[:8])}')
            except ImportError:
                messagebox.showwarning('Missing', 'Run: pip install openpyxl', parent=self)
            except Exception as ex:
                messagebox.showerror('Error', str(ex), parent=self)

        ctk.CTkButton(c, text='📂  Browse File', width=200, height=38,
                      fg_color=C['primary'], hover_color=C['primary_hover'],
                      command=_pick).pack()
        lbl = ctk.CTkLabel(c, text='No file selected.',
                           font=F['small'], text_color=C['text_secondary'],
                           wraplength=600)
        lbl.pack(pady=12)

    def _step1(self):
        c = self._content
        ctk.CTkLabel(c, text='Map your file columns to system fields:',
                     font=F['body'], text_color=C['text_primary']).pack(pady=(16, 12))
        if not self._headers:
            self._empty_msg(c, 'No file loaded. Go back to Step 1.'); return

        SYS_FIELDS = ['username', 'full_name', 'email', 'password',
                      'auth_type', 'description', 'tenant', '(ignore)']
        self._map_vars = {}
        for h in self._headers:
            row = ctk.CTkFrame(c, fg_color='transparent', height=32)
            row.pack(fill='x', pady=2)
            row.pack_propagate(False)
            ctk.CTkLabel(row, text=h, width=200, anchor='w',
                         font=F['small'],
                         text_color=C['text_primary']).pack(side='left')
            ctk.CTkLabel(row, text='→', width=24,
                         text_color=C['text_secondary']).pack(side='left')
            # Auto-guess mapping
            guess = '(ignore)'
            hl = h.lower().replace(' ', '_').replace('-', '_')
            for sf in SYS_FIELDS[:-1]:
                if hl in (sf, sf.replace('_', ''), sf[:4]):
                    guess = sf; break
            if 'name' in hl and 'full' not in hl: guess = 'username'
            if 'full' in hl:  guess = 'full_name'
            if 'mail' in hl:  guess = 'email'
            if 'pass' in hl:  guess = 'password'
            if 'auth' in hl:  guess = 'auth_type'

            var = ctk.StringVar(value=guess)
            self._map_vars[h] = var
            ctk.CTkOptionMenu(row, variable=var, values=SYS_FIELDS,
                              width=160,
                              fg_color=C['bg_tertiary'],
                              button_color=C['primary'],
                              dropdown_fg_color=C['bg_secondary'],
                              text_color=C['text_primary']).pack(side='left', padx=8)

    def _step2(self):
        self._mapping = {h: v.get() for h, v in self._map_vars.items()
                        } if hasattr(self, '_map_vars') else {}
        c = self._content
        ctk.CTkLabel(c, text='Validating…',
                     font=F['body'], text_color=C['text_primary']).pack(pady=(20, 8))

        errors = []; warnings = []; ok = []
        for i, row in enumerate(self._file_data, 1):
            rec = {sf: row.get(h, '') for h, sf in self._mapping.items()
                   if sf != '(ignore)'}
            uname = str(rec.get('username', '')).strip()
            email = str(rec.get('email', '')).strip()
            full  = str(rec.get('full_name', '')).strip()
            if not uname:
                errors.append(f'Row {i}: missing username'); continue
            if email and '@' not in email:
                warnings.append(f'Row {i}: invalid email "{email}"')
            if not full:
                warnings.append(f'Row {i}: missing full name for "{uname}"')
            ok.append(rec)

        self._validated = ok
        s = ctk.CTkFrame(c, fg_color=C['bg_secondary'], corner_radius=8)
        s.pack(fill='x', pady=8, ipady=8)
        ctk.CTkLabel(s, text=f'✅  Valid rows:    {len(ok)}',
                     font=F['small'], text_color=C['success'],
                     anchor='w').pack(fill='x', padx=16, pady=2)
        ctk.CTkLabel(s, text=f'⚠   Warnings:     {len(warnings)}',
                     font=F['small'], text_color=C['warning'],
                     anchor='w').pack(fill='x', padx=16, pady=2)
        ctk.CTkLabel(s, text=f'❌  Errors:        {len(errors)}',
                     font=F['small'], text_color=C['danger'],
                     anchor='w').pack(fill='x', padx=16, pady=2)
        for w in warnings[:10]:
            ctk.CTkLabel(c, text=f'  ⚠  {w}', font=F['small'],
                         text_color=C['warning'], anchor='w').pack(fill='x', padx=16)
        for e in errors[:10]:
            ctk.CTkLabel(c, text=f'  ❌  {e}', font=F['small'],
                         text_color=C['danger'],  anchor='w').pack(fill='x', padx=16)

    def _step3(self):
        c = self._content
        ctk.CTkLabel(c, text=f'Preview — {len(self._validated)} users to create:',
                     font=F['body'], text_color=C['text_primary']).pack(pady=(16, 8))
        for rec in self._validated[:20]:
            row = ctk.CTkFrame(c, fg_color=C['bg_secondary'],
                               corner_radius=6, height=34)
            row.pack(fill='x', pady=1)
            row.pack_propagate(False)
            ctk.CTkLabel(row, text=f"+ {rec.get('username','?')}",
                         width=180, anchor='w', font=F['small'],
                         text_color=C['success']).pack(side='left', padx=12)
            ctk.CTkLabel(row, text=rec.get('full_name', '—'),
                         anchor='w', font=F['small'],
                         text_color=C['text_secondary']).pack(side='left')
        if len(self._validated) > 20:
            ctk.CTkLabel(c, text=f'… and {len(self._validated)-20} more',
                         font=F['small'],
                         text_color=C['text_secondary']).pack(pady=4)

    def _step4(self): self._step3()  # same preview on Apply step

    def _apply(self):
        if not self._validated:
            messagebox.showwarning('Nothing to import', 'No valid rows.', parent=self); return
        ok = err = 0
        for rec in self._validated:
            try:
                success, _ = bo_session.create_user(
                    name=rec.get('username', ''),
                    password=rec.get('password', 'Welcome1!'),
                    email=rec.get('email', ''),
                    full_name=rec.get('full_name', ''),
                    auth_type=rec.get('auth_type', 'secEnterprise'),
                )
                if success: ok += 1
                else:       err += 1
            except Exception:
                err += 1
        messagebox.showinfo('Import Complete',
                            f'Created: {ok}  |  Errors: {err}', parent=self)
        if self._on_done:
            self._on_done()
        self.destroy()

    def _empty_msg(self, parent, text):
        ctk.CTkLabel(parent, text=text, font=F['small'],
                     text_color=C['text_secondary']).pack(pady=20)


# ═════════════════════════════════════════════════════════════════════════════
# GROUP DETAIL DRAWER
# ═════════════════════════════════════════════════════════════════════════════

class _GroupDrawer(ctk.CTkToplevel):
    TABS = ['Properties', 'Members', 'Member Of', 'Security']

    def __init__(self, master):
        super().__init__(master)
        self.title('Group Details')
        self.geometry('500x580')
        self.configure(fg_color=C['bg_primary'])
        self._gid = None
        self._build()

    def _build(self):
        tb = ctk.CTkFrame(self, fg_color=C['bg_tertiary'],
                          corner_radius=0, height=48)
        tb.pack(fill='x')
        tb.pack_propagate(False)
        self._title = ctk.CTkLabel(tb, text='', font=('Segoe UI', 14, 'bold'),
                                   text_color=C['text_primary'], anchor='w')
        self._title.pack(side='left', padx=16, fill='x', expand=True)
        ctk.CTkButton(tb, text='✕', width=32, height=28,
                      fg_color='transparent', hover_color=C['danger'],
                      text_color=C['text_secondary'],
                      command=self.withdraw).pack(side='right', padx=4)

        tab_bar = ctk.CTkFrame(self, fg_color=C['bg_secondary'],
                               corner_radius=0, height=36)
        tab_bar.pack(fill='x')
        tab_bar.pack_propagate(False)
        self._tab_btns   = {}
        self._tab_frames = {}
        self._active_tab = ctk.StringVar(value='Properties')
        for name in self.TABS:
            b = ctk.CTkButton(tab_bar, text=name, width=115, height=36,
                              font=('Segoe UI', 10),
                              fg_color='transparent',
                              hover_color=C['bg_tertiary'],
                              text_color=C['text_secondary'],
                              corner_radius=0,
                              command=lambda n=name: self._switch(n))
            b.pack(side='left')
            self._tab_btns[name] = b
            self._tab_frames[name] = ctk.CTkScrollableFrame(
                self, fg_color=C['bg_primary'], corner_radius=0)
        self._switch('Properties')

    def _switch(self, name):
        self._active_tab.set(name)
        for n, f in self._tab_frames.items():
            f.pack_forget()
        self._tab_frames[name].pack(fill='both', expand=True)
        for n, b in self._tab_btns.items():
            b.configure(fg_color=C['primary'] if n == name else 'transparent',
                        text_color='white' if n == name else C['text_secondary'])
        if self._gid:
            self._load(name)

    def load(self, gid, name):
        self._gid = gid
        self._title.configure(text=f'👥  {name}')
        self.deiconify(); self.lift()
        self._switch(self._active_tab.get())

    def _load(self, name):
        for w in self._tab_frames[name].winfo_children():
            w.destroy()
        gid = self._gid
        if name == 'Properties':
            _run_bg(lambda: bo_session.get_group_properties(gid), self._r_props)
        elif name == 'Members':
            _run_bg(lambda: bo_session.get_group_members(gid), self._r_members)
        elif name == 'Member Of':
            _run_bg(lambda: bo_session.get_group_member_of(gid), self._r_member_of)
        elif name == 'Security':
            _run_bg(lambda: bo_session.get_group_security(gid), self._r_security)

    def _r_props(self, d):
        f = self._tab_frames['Properties']
        if not d: ctk.CTkLabel(f, text='No data.', font=F['small'],
                               text_color=C['text_secondary']).pack(pady=20); return
        for label, key in [('Group Name', 'SI_NAME'), ('Description', 'SI_DESCRIPTION'),
                            ('Date Created', 'SI_CREATION_TIME'),
                            ('Date Modified', 'SI_UPDATE_TS')]:
            row = ctk.CTkFrame(f, fg_color='transparent')
            row.pack(fill='x', padx=16, pady=4)
            ctk.CTkLabel(row, text=label, width=155, anchor='nw',
                         font=('Segoe UI', 11),
                         text_color=C['text_secondary']).pack(side='left')
            val = _fmt(d.get(key, '')) if 'TIME' in key or 'TS' in key else d.get(key, '—')
            ctk.CTkLabel(row, text=str(val) or '—', anchor='nw',
                         font=('Segoe UI', 11),
                         wraplength=280,
                         text_color=C['text_primary']).pack(side='left', fill='x', expand=True)

    def _r_members(self, members):
        f = self._tab_frames['Members']
        if not members:
            ctk.CTkLabel(f, text='No members.', font=F['small'],
                         text_color=C['text_secondary']).pack(pady=20); return
        users  = [m for m in members if m.get('kind') == 'User']
        groups = [m for m in members if m.get('kind') == 'Group']
        if users:
            ctk.CTkLabel(f, text=f'Users ({len(users)})',
                         font=('Segoe UI', 11, 'bold'),
                         text_color=C['primary']).pack(anchor='w', padx=16, pady=(12, 4))
            for u in users:
                row = ctk.CTkFrame(f, fg_color=C['bg_secondary'],
                                   corner_radius=6, height=34)
                row.pack(fill='x', padx=16, pady=1)
                row.pack_propagate(False)
                ctk.CTkLabel(row, text='👤', width=26,
                             font=('Segoe UI', 13)).pack(side='left', padx=(10, 4))
                ctk.CTkLabel(row, text=u.get('name', '?'), anchor='w',
                             font=('Segoe UI', 11),
                             text_color=C['text_primary']).pack(side='left')
        if groups:
            ctk.CTkLabel(f, text=f'Sub-Groups ({len(groups)})',
                         font=('Segoe UI', 11, 'bold'),
                         text_color=C['primary']).pack(anchor='w', padx=16, pady=(12, 4))
            for g in groups:
                row = ctk.CTkFrame(f, fg_color=C['bg_secondary'],
                                   corner_radius=6, height=34)
                row.pack(fill='x', padx=16, pady=1)
                row.pack_propagate(False)
                ctk.CTkLabel(row, text='👥', width=26,
                             font=('Segoe UI', 13)).pack(side='left', padx=(10, 4))
                ctk.CTkLabel(row, text=g.get('name', '?'), anchor='w',
                             font=('Segoe UI', 11),
                             text_color=C['text_primary']).pack(side='left')

    def _r_member_of(self, groups):
        f = self._tab_frames['Member Of']
        if not groups:
            ctk.CTkLabel(f, text='Not a member of any group.', font=F['small'],
                         text_color=C['text_secondary']).pack(pady=20); return
        for g in groups:
            row = ctk.CTkFrame(f, fg_color=C['bg_secondary'],
                               corner_radius=6, height=34)
            row.pack(fill='x', padx=16, pady=2)
            row.pack_propagate(False)
            ctk.CTkLabel(row, text='👥', width=26,
                         font=('Segoe UI', 13)).pack(side='left', padx=(10, 4))
            ctk.CTkLabel(row, text=g.get('name', '?'), anchor='w',
                         font=('Segoe UI', 11),
                         text_color=C['text_primary']).pack(side='left')

    def _r_security(self, rights):
        f = self._tab_frames['Security']
        ctk.CTkLabel(f,
                     text='CMC → Users and Groups → [Group] → Security.',
                     font=('Segoe UI', 10),
                     text_color=C['text_secondary'],
                     wraplength=440).pack(pady=20, padx=16, anchor='w')


# ═════════════════════════════════════════════════════════════════════════════
# MAIN PAGE
# ═════════════════════════════════════════════════════════════════════════════

class UsersPage(ctk.CTkFrame):

    _USER_COLS = [
        # (id, heading, width, stretch)
        ('ck',      '',                    28,  False),   # checkbox placeholder
        ('name',    'Username',            155, True),
        ('full',    'Full Name',           170, True),
        ('email',   'Email',               180, True),
        ('auth',    'Auth Type',           110, False),
        ('status',  'Status',               90, False),
        ('risk',    'Risk',                110, False),
        ('tenant',  'Tenant',               80, False),
        ('created', 'Created',             125, False),
        ('modified','Last Modified',       125, False),
    ]
    _GROUP_COLS = [
        ('name',    'Group Name',          210, True),
        ('desc',    'Description',         310, True),
        ('created', 'Date Created',        140, False),
        ('modified','Date Modified',       140, False),
    ]

    def __init__(self, parent, **kw):
        super().__init__(parent, fg_color=C['bg_primary'], corner_radius=0, **kw)
        _PAGE_REF[0] = self.winfo_toplevel()
        self._all_users    = []
        self._all_groups   = []
        self._hier_data    = {}
        self._selected_ids = set()
        self._user_drawer  = None
        self._group_drawer = None
        self._build_ui()
        self._load_data()

    # ── UI construction ───────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Header row ────────────────────────────────────────────────────────
        hdr = ctk.CTkFrame(self, fg_color=C['bg_primary'],
                           corner_radius=0, height=52)
        hdr.pack(fill='x')
        hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text='User Management',
                     font=F['sub_header'],
                     text_color=C['text_primary']).pack(side='left', padx=20)
        _b = dict(height=32, corner_radius=6, font=F['small'])
        ctk.CTkButton(hdr, text='↓ Export', width=90,
                      fg_color=C['success'], hover_color=C['accent'],
                      command=self._export_dialog, **_b).pack(side='right', padx=4, pady=10)
        ctk.CTkButton(hdr, text='↑ Import', width=90,
                      fg_color=C['secondary'],
                      hover_color='#7C3AED',
                      command=self._open_import, **_b).pack(side='right', padx=2, pady=10)
        ctk.CTkButton(hdr, text='⟳ Refresh', width=90,
                      fg_color=C['bg_tertiary'],
                      hover_color=C['bg_secondary'],
                      command=self._refresh, **_b).pack(side='right', padx=4, pady=10)

        # ── Insight cards ─────────────────────────────────────────────────────
        cards_bar = ctk.CTkFrame(self, fg_color=C['bg_primary'],
                                 corner_radius=0, height=82)
        cards_bar.pack(fill='x', padx=10)
        cards_bar.pack_propagate(False)
        self._c_total    = _card(cards_bar, 'Total Users',      '—', C['primary'])
        self._c_active   = _card(cards_bar, 'Active',           '—', C['success'])
        self._c_disabled = _card(cards_bar, 'Disabled',         '—', C['danger'])
        self._c_locked   = _card(cards_bar, 'Locked',           '—', C['warning'])
        self._c_external = _card(cards_bar, 'External Auth',    '—', C['secondary'])
        self._c_never    = _card(cards_bar, 'Never Logged In',  '—', '#F59E0B')

        # ── Section tabs ──────────────────────────────────────────────────────
        tab_bar = ctk.CTkFrame(self, fg_color=C['bg_secondary'],
                               corner_radius=0, height=40)
        tab_bar.pack(fill='x')
        tab_bar.pack_propagate(False)
        self._tab_btns  = {}
        self._active_sec = ctk.StringVar(value='Users')
        for name in ('Users', 'Groups', 'Hierarchy'):
            b = ctk.CTkButton(
                tab_bar, text=name, width=110, height=40,
                font=F['body'],
                fg_color=C['primary'] if name == 'Users' else 'transparent',
                hover_color=C['bg_tertiary'],
                text_color='white' if name == 'Users' else C['text_secondary'],
                corner_radius=0,
                command=lambda n=name: self._switch_sec(n))
            b.pack(side='left')
            self._tab_btns[name] = b

        # ── Search + Filter row ───────────────────────────────────────────────
        sf = ctk.CTkFrame(self, fg_color=C['bg_secondary'],
                          corner_radius=0, height=46)
        sf.pack(fill='x')
        sf.pack_propagate(False)
        self._search_var = ctk.StringVar()
        self._search_var.trace_add('write', lambda *_: self._apply_filter())
        ctk.CTkEntry(sf, placeholder_text='🔍  Search by name, full name, email…',
                     textvariable=self._search_var,
                     width=320, height=30,
                     fg_color=C['bg_tertiary'],
                     border_color=C['bg_tertiary'],
                     text_color=C['text_primary'],
                     font=F['small']).pack(side='left', padx=10, pady=8)

        ctk.CTkLabel(sf, text='Status:',
                     font=F['small'],
                     text_color=C['text_secondary']).pack(side='left', padx=(8, 2))
        self._filter_status = ctk.StringVar(value='All')
        ctk.CTkOptionMenu(sf, variable=self._filter_status,
                          values=['All', 'Enabled', 'Disabled', 'Locked'],
                          width=110, height=28,
                          command=lambda _: self._apply_filter(),
                          fg_color=C['bg_tertiary'],
                          button_color=C['primary'],
                          dropdown_fg_color=C['bg_secondary'],
                          text_color=C['text_primary'],
                          font=F['small']).pack(side='left', padx=2)

        ctk.CTkLabel(sf, text='Auth:',
                     font=F['small'],
                     text_color=C['text_secondary']).pack(side='left', padx=(10, 2))
        self._filter_auth = ctk.StringVar(value='All')
        ctk.CTkOptionMenu(sf, variable=self._filter_auth,
                          values=['All', 'Enterprise', 'LDAP', 'Windows AD', 'SAP'],
                          width=120, height=28,
                          command=lambda _: self._apply_filter(),
                          fg_color=C['bg_tertiary'],
                          button_color=C['primary'],
                          dropdown_fg_color=C['bg_secondary'],
                          text_color=C['text_primary'],
                          font=F['small']).pack(side='left', padx=2)

        self._status_lbl = ctk.StringVar(value='')
        ctk.CTkLabel(sf, textvariable=self._status_lbl,
                     font=F['small'],
                     text_color=C['text_secondary']).pack(side='right', padx=12)

        # ── Bulk actions toolbar (hidden until rows selected) ──────────────────
        self._bulk_bar = ctk.CTkFrame(self, fg_color=C['bg_tertiary'],
                                      corner_radius=0, height=40)
        # NOT packed yet — shown only when selection > 0
        self._bulk_bar.pack_propagate(False)
        ctk.CTkLabel(self._bulk_bar, text='',
                     textvariable=ctk.StringVar(value=''),
                     font=F['small'],
                     text_color=C['text_primary']).pack(side='left', padx=10)
        self._sel_lbl = ctk.CTkLabel(self._bulk_bar, text='0 selected',
                                     font=F['small'],
                                     text_color=C['text_primary'])
        self._sel_lbl.pack(side='left', padx=10)
        _ba = dict(height=28, corner_radius=6, font=F['small'], width=110)
        ctk.CTkButton(self._bulk_bar, text='✓ Enable',
                      fg_color=C['success'], hover_color=C['accent'],
                      command=lambda: self._bulk_action('enable'), **_ba
                      ).pack(side='left', padx=3)
        ctk.CTkButton(self._bulk_bar, text='✗ Disable',
                      fg_color=C['danger'], hover_color='#DC2626',
                      command=lambda: self._bulk_action('disable'), **_ba
                      ).pack(side='left', padx=3)
        ctk.CTkButton(self._bulk_bar, text='⟳ Reset PWD',
                      fg_color=C['warning'], hover_color='#D97706',
                      command=lambda: self._bulk_action('reset'), **_ba
                      ).pack(side='left', padx=3)
        ctk.CTkButton(self._bulk_bar, text='↓ Export Sel.',
                      fg_color=C['primary'], hover_color=C['primary_hover'],
                      command=self._export_selected, **_ba
                      ).pack(side='left', padx=3)
        ctk.CTkButton(self._bulk_bar, text='✕ Clear',
                      fg_color='transparent',
                      hover_color=C['bg_secondary'],
                      text_color=C['text_secondary'],
                      command=self._clear_selection, **_ba
                      ).pack(side='right', padx=8)

        # ── Main list frames ──────────────────────────────────────────────────
        self._list_area = ctk.CTkFrame(self, fg_color=C['bg_primary'], corner_radius=0)
        self._list_area.pack(fill='both', expand=True)

        # Users tree
        self._user_frame = ctk.CTkFrame(self._list_area,
                                         fg_color=C['bg_secondary'], corner_radius=0)
        self._user_tv = _tree(self._user_frame, self._USER_COLS, row_h=34)
        # Groups tree
        self._group_frame = ctk.CTkFrame(self._list_area,
                                          fg_color=C['bg_secondary'], corner_radius=0)
        self._group_tv = _tree(self._group_frame, self._GROUP_COLS, row_h=34)
        # Hierarchy
        self._hier_frame = ctk.CTkScrollableFrame(self._list_area,
                                                   fg_color=C['bg_primary'],
                                                   corner_radius=0)
        # Tag colors for status
        for tag, color in [('Enabled', C['text_primary']),
                            ('Disabled', C['danger']),
                            ('Locked', C['warning'])]:
            self._user_tv.tag_configure(tag, foreground=color)

        # Bindings
        self._user_tv.bind('<ButtonRelease-1>',   self._on_user_click)
        self._user_tv.bind('<<TreeviewSelect>>', self._on_user_select)
        self._group_tv.bind('<ButtonRelease-1>',  self._on_group_click)

        self._switch_sec('Users')

    # ── Section switching ─────────────────────────────────────────────────────

    def _switch_sec(self, name):
        self._active_sec.set(name)
        self._user_frame.pack_forget()
        self._group_frame.pack_forget()
        self._hier_frame.pack_forget()
        if name == 'Users':
            self._user_frame.pack(fill='both', expand=True)
        elif name == 'Groups':
            self._group_frame.pack(fill='both', expand=True)
        else:
            self._hier_frame.pack(fill='both', expand=True)
            self._render_hierarchy()
        for n, b in self._tab_btns.items():
            b.configure(fg_color=C['primary'] if n == name else 'transparent',
                        text_color='white' if n == name else C['text_secondary'])

    # ── Data loading ──────────────────────────────────────────────────────────

    def _load_data(self):
        self._status_lbl.set('⏳ Loading…')
        _run_bg(self._fetch, self._populate)

    def _fetch(self):
        users  = bo_session.get_users_detailed()  or []
        groups = bo_session.get_groups_detailed() or []
        hier   = bo_session.get_hierarchy_data()  or {}
        return users, groups, hier

    def _populate(self, result):
        try:
            if not self.winfo_exists():
                return
        except Exception:
            return
        if not result:
            self._status_lbl.set('❌ Failed to load.'); return
        users, groups, hier = result
        self._all_users  = users
        self._all_groups = groups
        self._hier_data  = hier
        try:
            self._update_cards(users)
        except Exception:
            pass
        self._fill_users(users)
        self._fill_groups(groups)
        self._status_lbl.set(f'{len(users)} users  |  {len(groups)} groups')
        if self._active_sec.get() == 'Hierarchy':
            self._render_hierarchy()

    def _update_cards(self, users):
        total    = len(users)
        active   = sum(1 for u in users if u.get('account_status') == 'Enabled')
        disabled = sum(1 for u in users if u.get('account_status') == 'Disabled')
        locked   = sum(1 for u in users if u.get('account_status') == 'Locked')
        external = sum(1 for u in users if u.get('auth_type', 'Enterprise') != 'Enterprise')
        never    = sum(1 for u in users if not u.get('date_modified'))
        self._c_total.configure(text=str(total))
        self._c_active.configure(text=str(active))
        self._c_disabled.configure(text=str(disabled))
        self._c_locked.configure(text=str(locked))
        self._c_external.configure(text=str(external))
        self._c_never.configure(text=str(never))

    def _fill_users(self, users):
        try:
            tv = self._user_tv
            if not tv.winfo_exists():
                return
            for r in tv.get_children():
                tv.delete(r)
        except Exception:
            return
        tv.tag_configure('Enabled',  foreground=C['text_primary'])
        tv.tag_configure('Disabled', foreground=C['danger'])
        tv.tag_configure('Locked',   foreground=C['warning'])
        for u in users:
            status = u.get('account_status', 'Enabled')
            risks  = _risk_tags(u)
            risk_s = ' '.join(f'⚠{r[:4]}' for r in risks) if risks else '✓'
            try:
                tv.insert('', 'end', iid=str(u['id']),
                          values=('',
                                  u.get('name', ''),
                                  u.get('full_name', ''),
                                  u.get('email', ''),
                                  u.get('auth_type', ''),
                                  status,
                                  risk_s,
                                  u.get('tenant', ''),
                                  _fmt(u.get('date_created', '')),
                                  _fmt(u.get('date_modified', ''))),
                          tags=(status,))
            except Exception:
                pass

    def _fill_groups(self, groups):
        tv = self._group_tv
        for r in tv.get_children():
            tv.delete(r)
        for g in groups:
            tv.insert('', 'end', iid=str(g['id']),
                      values=(g.get('name', ''),
                              g.get('description', ''),
                              _fmt(g.get('date_created', '')),
                              _fmt(g.get('date_modified', ''))))

    # ── Filter ────────────────────────────────────────────────────────────────

    def _apply_filter(self):
        q      = self._search_var.get().lower().strip()
        st     = self._filter_status.get()
        auth   = self._filter_auth.get()
        shown  = 0
        tv = self._user_tv
        for iid in tv.get_children():
            vals = tv.item(iid, 'values')
            # vals: ('', name, full, email, auth, status, risk, tenant, created, modified)
            name_full = f'{vals[1]} {vals[2]} {vals[3]}'.lower()
            match_q    = not q    or q in name_full
            match_st   = st   == 'All' or vals[5] == st
            match_auth = auth == 'All' or vals[4] == auth
            if match_q and match_st and match_auth:
                tv.reattach(iid, '', 'end')
                shown += 1
            else:
                tv.detach(iid)
        self._status_lbl.set(f'Showing {shown} of {len(self._all_users)} users')

    # ── Click / selection handlers ────────────────────────────────────────────

    def _on_user_click(self, event):
        sel = self._user_tv.selection()
        if not sel:
            return
        uid = sel[0]
        u   = next((u for u in self._all_users if str(u['id']) == uid), None)
        if not u:
            return
        if self._user_drawer is None or not self._user_drawer.winfo_exists():
            self._user_drawer = _UserDrawer(self)
        self._user_drawer.load(u)

    def _on_user_select(self, event):
        sel = set(self._user_tv.selection())
        if sel != self._selected_ids:
            self._selected_ids = sel
            self._sel_lbl.configure(text=f'{len(sel)} selected')
            if sel:
                self._bulk_bar.pack(fill='x', before=self._list_area)
            else:
                self._bulk_bar.pack_forget()

    def _on_group_click(self, event):
        sel = self._group_tv.selection()
        if not sel:
            return
        gid  = sel[0]
        g    = next((g for g in self._all_groups if str(g['id']) == gid), None)
        if not g:
            return
        if self._group_drawer is None or not self._group_drawer.winfo_exists():
            self._group_drawer = _GroupDrawer(self)
        self._group_drawer.load(int(gid), g.get('name', ''))

    def _clear_selection(self):
        self._user_tv.selection_remove(*self._user_tv.selection())
        self._selected_ids.clear()
        self._bulk_bar.pack_forget()

    # ── Bulk actions ──────────────────────────────────────────────────────────

    def _bulk_action(self, action):
        ids = list(self._selected_ids)
        if not ids:
            return
        label = {'enable': 'enable', 'disable': 'disable',
                 'reset': 'reset password for'}[action]
        if not messagebox.askyesno('Confirm',
                                   f'Are you sure you want to {label} '
                                   f'{len(ids)} user(s)?'):
            return
        def _do():
            ok = err = 0
            for uid in ids:
                try:
                    if action == 'enable':
                        s, _ = bo_session.disable_user(int(uid), False)
                    elif action == 'disable':
                        s, _ = bo_session.disable_user(int(uid), True)
                    else:
                        s, _ = bo_session.reset_user_password(int(uid), 'Welcome1!')
                    if s: ok += 1
                    else: err += 1
                except Exception:
                    err += 1
            self.after(0, lambda: (
                messagebox.showinfo('Done', f'OK: {ok}  Errors: {err}'),
                self._refresh()
            ))
        threading.Thread(target=_do, daemon=True).start()

    # ── Hierarchy view ────────────────────────────────────────────────────────

    def _render_hierarchy(self):
        f = self._hier_frame
        for w in f.winfo_children():
            w.destroy()
        hier = self._hier_data
        if not hier:
            ctk.CTkLabel(f, text='Click Refresh to load hierarchy.',
                         font=F['small'],
                         text_color=C['text_secondary']).pack(pady=24)
            return

        ctk.CTkLabel(f, text=f'Group Hierarchy  —  {len(hier)} group(s)',
                     font=F['body'], text_color=C['text_primary'],
                     anchor='w').pack(fill='x', padx=16, pady=(12, 6))

        for gname, gdata in sorted(hier.items()):
            members = gdata.get('members', []) if isinstance(gdata, dict) else []
            card = ctk.CTkFrame(f, fg_color=C['bg_secondary'], corner_radius=8)
            card.pack(fill='x', padx=14, pady=4)
            top = ctk.CTkFrame(card, fg_color='transparent', height=42)
            top.pack(fill='x')
            top.pack_propagate(False)
            ctk.CTkLabel(top, text='👥', width=32,
                         font=('Segoe UI', 15)).pack(side='left', padx=(10, 4))
            ctk.CTkLabel(top, text=gname, font=('Segoe UI', 12, 'bold'),
                         text_color=C['primary'], anchor='w').pack(
                             side='left', fill='x', expand=True)
            ctk.CTkLabel(top, text=f'{len(members)} member(s)',
                         font=F['small'],
                         text_color=C['text_secondary']).pack(side='right', padx=12)
            for m in members[:8]:
                mrow = ctk.CTkFrame(card, fg_color='transparent', height=26)
                mrow.pack(fill='x')
                mrow.pack_propagate(False)
                ctk.CTkLabel(mrow, text='   👤',
                             width=36, font=('Segoe UI', 11)).pack(side='left', padx=(16, 2))
                ctk.CTkLabel(mrow, text=str(m), anchor='w',
                             font=('Segoe UI', 10),
                             text_color=C['text_secondary']).pack(side='left')
            if len(members) > 8:
                ctk.CTkLabel(card,
                             text=f'  … and {len(members)-8} more',
                             font=F['small'],
                             text_color=C['text_secondary']).pack(
                                 anchor='w', padx=28, pady=(0, 6))

    # ── Export dialog ─────────────────────────────────────────────────────────

    def _export_dialog(self):
        dlg = ctk.CTkToplevel(self)
        dlg.title('Export Users')
        dlg.geometry('460x420')
        dlg.configure(fg_color=C['bg_primary'])
        dlg.attributes('-topmost', True)

        ctk.CTkLabel(dlg, text='Export Options',
                     font=('Segoe UI', 16, 'bold'),
                     text_color=C['text_primary']).pack(pady=(20, 4))

        # Scope
        ctk.CTkLabel(dlg, text='Scope:', font=F['body'],
                     text_color=C['text_secondary'],
                     anchor='w').pack(fill='x', padx=24, pady=(12, 4))
        scope_var = ctk.StringVar(value='all')
        for val, lbl in [('all', 'All Users'), ('filtered', 'Filtered / Visible Users'),
                          ('selected', f'Selected Users ({len(self._selected_ids)})')]:
            ctk.CTkRadioButton(dlg, text=lbl, variable=scope_var, value=val,
                               font=F['small'],
                               text_color=C['text_primary'],
                               fg_color=C['primary']).pack(anchor='w', padx=40, pady=2)

        # Options
        ctk.CTkLabel(dlg, text='Include:', font=F['body'],
                     text_color=C['text_secondary'],
                     anchor='w').pack(fill='x', padx=24, pady=(12, 4))
        inc_groups = ctk.BooleanVar(value=True)
        inc_hier   = ctk.BooleanVar(value=False)
        ctk.CTkCheckBox(dlg, text='Group memberships (fetches live)',
                        variable=inc_groups, font=F['small'],
                        text_color=C['text_primary'],
                        fg_color=C['primary']).pack(anchor='w', padx=40, pady=2)
        ctk.CTkCheckBox(dlg, text='Hierarchy summary',
                        variable=inc_hier, font=F['small'],
                        text_color=C['text_primary'],
                        fg_color=C['primary']).pack(anchor='w', padx=40, pady=2)

        # Format
        ctk.CTkLabel(dlg, text='Format:', font=F['body'],
                     text_color=C['text_secondary'],
                     anchor='w').pack(fill='x', padx=24, pady=(12, 4))
        fmt_var = ctk.StringVar(value='excel')
        for val, lbl in [('excel', 'Excel (.xlsx) with charts'),
                          ('csv',   'CSV (.csv)'),
                          ('json',  'JSON (.json)')]:
            ctk.CTkRadioButton(dlg, text=lbl, variable=fmt_var, value=val,
                               font=F['small'],
                               text_color=C['text_primary'],
                               fg_color=C['primary']).pack(anchor='w', padx=40, pady=2)

        def _do_export():
            scope = scope_var.get()
            fmt   = fmt_var.get()
            if scope == 'all':
                data = self._all_users
            elif scope == 'selected':
                data = [u for u in self._all_users
                        if str(u['id']) in self._selected_ids]
            else:
                data = [u for u in self._all_users
                        if self._user_tv.exists(str(u['id']))]
            dlg.destroy()
            self._do_export(data, fmt, inc_groups.get(), inc_hier.get())

        ctk.CTkButton(dlg, text='Export', width=140, height=36,
                      fg_color=C['success'], hover_color=C['accent'],
                      command=_do_export,
                      font=('Segoe UI', 13, 'bold')).pack(pady=20)

    def _do_export(self, users, fmt, inc_groups, inc_hier):
        exts = {'excel': '.xlsx', 'csv': '.csv', 'json': '.json'}
        path = filedialog.asksaveasfilename(
            defaultextension=exts[fmt],
            filetypes=[('All', f'*{exts[fmt]}')],
            initialfile=f'bo_users{exts[fmt]}')
        if not path:
            return

        # Build flat records
        records = []
        for u in users:
            rec = {
                'Username':     u.get('name', ''),
                'Full Name':    u.get('full_name', ''),
                'Email':        u.get('email', ''),
                'Description':  u.get('description', ''),
                'Status':       u.get('account_status', ''),
                'Auth Type':    u.get('auth_type', ''),
                'Tenant':       u.get('tenant', ''),
                'Date Created': _fmt(u.get('date_created', '')),
                'Last Modified':_fmt(u.get('date_modified', '')),
                'Risk Flags':   ', '.join(_risk_tags(u)),
            }
            if inc_groups:
                grps = bo_session.get_user_member_of(u['id'])
                rec['Groups'] = '; '.join(g.get('name', '') for g in grps)
            records.append(rec)

        if fmt == 'csv':
            with open(path, 'w', newline='', encoding='utf-8-sig') as fp:
                w = csv.DictWriter(fp, fieldnames=list(records[0].keys()))
                w.writeheader()
                w.writerows(records)
            messagebox.showinfo('Export', f'CSV saved:\n{path}')

        elif fmt == 'json':
            with open(path, 'w', encoding='utf-8') as fp:
                json.dump(records, fp, indent=2, ensure_ascii=False)
            messagebox.showinfo('Export', f'JSON saved:\n{path}')

        else:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.chart import BarChart, PieChart, Reference
                from openpyxl.utils import get_column_letter

                wb = Workbook()
                ws = wb.active
                ws.title = 'Users'

                HDR_FILL = PatternFill('solid', fgColor='1E3A5F')
                HDR_FONT = Font(color='FFFFFF', bold=True, name='Segoe UI', size=11)
                ALT_FILL = PatternFill('solid', fgColor='EBF2FF')
                THIN     = Border(
                    left=Side('thin', color='CCCCCC'),
                    right=Side('thin', color='CCCCCC'),
                    top=Side('thin', color='CCCCCC'),
                    bottom=Side('thin', color='CCCCCC'))

                cols = list(records[0].keys())
                # Title
                ws.merge_cells(f'A1:{get_column_letter(len(cols))}1')
                ws['A1'] = f'SAP BO Users Export  —  {len(records)} users'
                ws['A1'].font = Font(bold=True, size=14, name='Segoe UI',
                                     color='1E3A5F')
                ws['A1'].alignment = Alignment(horizontal='center')
                ws.row_dimensions[1].height = 26
                # Headers
                for ci, col in enumerate(cols, 1):
                    cell = ws.cell(row=2, column=ci, value=col)
                    cell.fill = HDR_FILL
                    cell.font = HDR_FONT
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = THIN
                ws.row_dimensions[2].height = 20
                # Data
                for ri, rec in enumerate(records, 3):
                    fill = ALT_FILL if ri % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
                    for ci, col in enumerate(cols, 1):
                        cell = ws.cell(row=ri, column=ci, value=rec[col])
                        cell.fill = fill
                        cell.border = THIN
                # Widths
                for ci, col in enumerate(cols, 1):
                    maxw = max(len(col), max((len(str(r[col])) for r in records), default=0))
                    ws.column_dimensions[get_column_letter(ci)].width = min(maxw + 4, 50)
                ws.freeze_panes = 'A3'

                # Sheet 2: Status pie chart
                ws2 = wb.create_sheet('📊 Status Chart')
                status_counts = {}
                for rec in records:
                    st = rec.get('Status', 'Unknown')
                    status_counts[st] = status_counts.get(st, 0) + 1
                ws2['A1'] = 'Status'; ws2['B1'] = 'Count'
                for i, (st, cnt) in enumerate(status_counts.items(), 2):
                    ws2[f'A{i}'] = st; ws2[f'B{i}'] = cnt
                pie = PieChart()
                pie.title = 'Users by Account Status'
                pie.style  = 10
                pie.width  = 20; pie.height = 14
                dr = Reference(ws2, min_col=2, min_row=1,
                                max_row=len(status_counts) + 1)
                cr = Reference(ws2, min_col=1, min_row=2,
                                max_row=len(status_counts) + 1)
                pie.add_data(dr, titles_from_data=True)
                pie.set_categories(cr)
                ws2.add_chart(pie, 'D2')

                # Sheet 3: Auth type bar chart
                ws3 = wb.create_sheet('🔐 Auth Chart')
                auth_counts = {}
                for rec in records:
                    at = rec.get('Auth Type', 'Unknown')
                    auth_counts[at] = auth_counts.get(at, 0) + 1
                ws3['A1'] = 'Auth Type'; ws3['B1'] = 'Count'
                for i, (at, cnt) in enumerate(auth_counts.items(), 2):
                    ws3[f'A{i}'] = at; ws3[f'B{i}'] = cnt
                bar = BarChart()
                bar.type  = 'col'
                bar.title = 'Users by Authentication Type'
                bar.style = 10
                bar.width = 22; bar.height = 14
                bar.y_axis.title = 'Count'
                bar.x_axis.title = 'Auth Type'
                dr2 = Reference(ws3, min_col=2, min_row=1,
                                 max_row=len(auth_counts) + 1)
                cr2 = Reference(ws3, min_col=1, min_row=2,
                                 max_row=len(auth_counts) + 1)
                bar.add_data(dr2, titles_from_data=True)
                bar.set_categories(cr2)
                bar.series[0].graphicalProperties.solidFill = '2563EB'
                ws3.add_chart(bar, 'D2')

                wb.save(path)
                messagebox.showinfo('Export Complete',
                                    f'Excel with charts saved:\n{path}\n\n'
                                    f'Sheets: Users data, Status Pie, Auth Bar')
            except ImportError:
                messagebox.showwarning('Missing', 'Run: pip install openpyxl')
            except Exception as e:
                messagebox.showerror('Export Error', str(e))

    def _export_selected(self):
        sel_users = [u for u in self._all_users
                     if str(u['id']) in self._selected_ids]
        if not sel_users:
            messagebox.showinfo('No Selection', 'Select rows first.'); return
        self._do_export(sel_users, 'excel', False, False)

    # ── Import ────────────────────────────────────────────────────────────────

    def _open_import(self):
        wizard = _ImportWizard(self, on_done=self._refresh)
        wizard.focus()

    # ── Toolbar ───────────────────────────────────────────────────────────────

    def _refresh(self):
        self._clear_selection()
        self._load_data()
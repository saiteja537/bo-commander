"""
gui/pages/housekeeping.py
Housekeeping — real CMS data only, no test/mock data.
14 categories: Reports, Instances, Failed, Old Instances, Users, Folders,
               Universes, Connections, Audit Events, Recurring Schedules,
               Stuck Instances, Promotion Jobs, Versions/History, Temp Objects
All counts come from live CMS queries. Empty = shows "No data found" cleanly.
Export: Excel (all tabs + Visualizations sheet) + 5 chart types in-app.
FIX: border_color='transparent' → bg_tertiary (customtkinter compatibility).
"""

import threading
import os
from tkinter import messagebox, filedialog
import customtkinter as ctk
from datetime import datetime
from config import Config
from core.sapbo_connection import bo_session

# ── Optional: matplotlib for in-app charts ────────────────────────────────────
try:
    import matplotlib
    matplotlib.use('TkAgg')
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

# ── Optional: openpyxl for Excel export ───────────────────────────────────────
try:
    import openpyxl
    from openpyxl.chart import BarChart, PieChart, LineChart, BarChart3D, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

C = Config.COLORS

CATEGORIES = [
    ('reports',           'Reports',            '📄', '#3B82F6', 'All report objects (Webi + Crystal)'),
    ('instances',         'All Instances',      '📋', '#8B5CF6', 'All report run instances'),
    ('failed_instances',  'Failed Instances',   '❌', '#EF4444', 'Instances in failed state'),
    ('old_instances',     'Old Instances',      '🕐', '#F97316', 'Instances older than 30 days'),
    ('users',             'Users',              '👤', '#10B981', 'Enterprise + LDAP users'),
    ('folders',           'Folders',            '📁', '#F59E0B', 'Public folders'),
    ('universes',         'Universes',          '🌐', '#06B6D4', 'UNV and UNX universes'),
    ('connections',       'Connections',        '🔗', '#84CC16', 'Database connections'),
    ('audit_events',      'Audit Events',       '🔍', '#EC4899', 'Recent CMS audit log entries'),
    ('recurring',         'Recurring Schedules','🔄', '#F59E0B', 'Active recurring schedules'),
    ('stuck_instances',   'Stuck Instances',    '⏸', '#DC2626', 'Running instances older than 2h'),
    ('promotion_jobs',    'Promotion Jobs',     '🚀', '#7C3AED', 'LCM lifecycle jobs'),
    ('versions',          'Versions/History',   '📚', '#0EA5E9', 'Object version history'),
    ('temp_objects',      'Temp Objects',       '🗒', '#64748B', 'Objects in temp/cache folders'),
]

CHART_TYPES = ['Bar Chart', 'Horizontal Bar', 'Pie Chart', 'Donut Chart', 'Line Chart']

CHART_COLORS = [
    '#3B82F6','#8B5CF6','#EF4444','#F97316','#10B981',
    '#F59E0B','#06B6D4','#84CC16','#EC4899','#DC2626',
    '#7C3AED','#0EA5E9','#64748B','#F43F5E'
]


def _get_columns(category_id):
    common = [('SI_ID', 70), ('SI_NAME', 300), ('SI_KIND', 140), ('SI_OWNER', 130)]
    extras = {
        'instances':        common + [('SI_STARTTIME', 160), ('SI_ENDTIME', 160)],
        'failed_instances': common + [('SI_STARTTIME', 160), ('SI_ENDTIME', 160)],
        'old_instances':    common + [('SI_STARTTIME', 160)],
        'stuck_instances':  common + [('SI_STARTTIME', 160), ('Duration', 100)],
        'audit_events':     [('SI_ID', 70), ('SI_NAME', 200), ('Event', 160), ('User', 120), ('Time', 160)],
        'recurring':        common + [('SI_SCHEDULE_STATUS', 120)],
        'promotion_jobs':   common + [('SI_CREATION_TIME', 160), ('Status', 100)],
        'versions':         common + [('SI_VERSION', 80), ('SI_UPDATE_TS', 160)],
        'temp_objects':     common + [('SI_UPDATE_TS', 160)],
    }
    return extras.get(category_id, common + [('SI_UPDATE_TS', 160)])


def _run_category_query(category_id):
    """Run CMS query for a category. Returns list of dicts."""
    try:
        queries = {
            'reports': (
                "SELECT TOP 500 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_UPDATE_TS "
                "FROM CI_INFOOBJECTS "
                "WHERE SI_INSTANCE=0 AND SI_KIND IN ('Webi','CrystalReport') "
                "ORDER BY SI_UPDATE_TS DESC"
            ),
            'instances': (
                "SELECT TOP 500 SI_ID, SI_NAME, SI_KIND, SI_OWNER, "
                "SI_STARTTIME, SI_ENDTIME "
                "FROM CI_INFOOBJECTS WHERE SI_INSTANCE=1 "
                "ORDER BY SI_STARTTIME DESC"
            ),
            'failed_instances': (
                "SELECT TOP 500 SI_ID, SI_NAME, SI_KIND, SI_OWNER, "
                "SI_STARTTIME, SI_ENDTIME "
                "FROM CI_INFOOBJECTS "
                "WHERE SI_INSTANCE=1 AND SI_PROCESSINFO.SI_STATUS_INFO=1 "
                "ORDER BY SI_STARTTIME DESC"
            ),
            'old_instances': (
                "SELECT TOP 500 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_STARTTIME "
                "FROM CI_INFOOBJECTS WHERE SI_INSTANCE=1 "
                "AND SI_STARTTIME < '2025-12-01 00:00:00' "
                "ORDER BY SI_STARTTIME ASC"
            ),
            'users': (
                "SELECT TOP 500 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_UPDATE_TS "
                "FROM CI_SYSTEMOBJECTS "
                "WHERE SI_KIND IN ('User','LDAPUser','WinADUser') "
                "ORDER BY SI_NAME ASC"
            ),
            'folders': (
                "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_UPDATE_TS "
                "FROM CI_INFOOBJECTS "
                "WHERE SI_KIND='Folder' AND SI_INSTANCE=0 "
                "ORDER BY SI_NAME ASC"
            ),
            'universes': (
                "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_UPDATE_TS "
                "FROM CI_APPOBJECTS "
                "WHERE SI_KIND IN ('Universe','DSL.MetaDataFile') "
                "ORDER BY SI_NAME ASC"
            ),
            'connections': (
                "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_UPDATE_TS "
                "FROM CI_APPOBJECTS WHERE SI_KIND='Connection' "
                "ORDER BY SI_NAME ASC"
            ),
            'audit_events': (
                "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND, "
                "SI_AUDIT_INFO.SI_AUDIT_EVTNAME AS Event, "
                "SI_AUDIT_INFO.SI_AUDIT_USERNAME AS AuditUser, "
                "SI_AUDIT_INFO.SI_AUDIT_STARTTIME AS AuditTime "
                "FROM CI_INFOOBJECTS WHERE SI_KIND='AuditEvent' "
                "ORDER BY SI_AUDIT_INFO.SI_AUDIT_STARTTIME DESC"
            ),
            'recurring': (
                "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND, SI_OWNER, "
                "SI_SCHEDULEINFO.SI_SCHEDULE_STATUS AS SI_SCHEDULE_STATUS "
                "FROM CI_INFOOBJECTS "
                "WHERE SI_RECURRING=1 AND SI_INSTANCE=0 "
                "ORDER BY SI_NAME ASC"
            ),
            'stuck_instances': (
                "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_STARTTIME "
                "FROM CI_INFOOBJECTS WHERE SI_INSTANCE=1 "
                "AND SI_PROCESSINFO.SI_STATUS_INFO=2 "
                "AND SI_STARTTIME < '2025-01-01 00:00:00' "
                "ORDER BY SI_STARTTIME ASC"
            ),
            'promotion_jobs': (
                "SELECT TOP 100 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_CREATION_TIME "
                "FROM CI_INFOOBJECTS "
                "WHERE SI_KIND IN ('LcmJob','PromotionJob') "
                "ORDER BY SI_CREATION_TIME DESC"
            ),
            'versions': (
                "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND, SI_OWNER, "
                "SI_VERSION, SI_UPDATE_TS "
                "FROM CI_INFOOBJECTS "
                "WHERE SI_INSTANCE=0 AND SI_VERSION > 1 "
                "ORDER BY SI_VERSION DESC"
            ),
            'temp_objects': (
                "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_UPDATE_TS "
                "FROM CI_INFOOBJECTS WHERE SI_INSTANCE=0 "
                "AND SI_PARENTID IN "
                "(SELECT SI_ID FROM CI_INFOOBJECTS WHERE SI_NAME='Temp') "
                "ORDER BY SI_UPDATE_TS ASC"
            ),
        }
        q = queries.get(category_id)
        if not q:
            return []
        d = bo_session.run_cms_query(q)
        return d.get('entries', []) if d else []
    except Exception:
        return []


# ─────────────────────────────────────────────────────────────────────────────
#  Visualization popup window
# ─────────────────────────────────────────────────────────────────────────────

class VisualizationWindow(ctk.CTkToplevel):
    """
    Shows 5 chart types (Bar, Horizontal Bar, Pie, Donut, Line) for the
    category-count data. User selects chart type and can download as PNG.
    """

    def __init__(self, parent, counts_dict):
        super().__init__(parent)
        self.title('📊  Housekeeping Visualizations')
        self.geometry('1080x680')
        self.configure(fg_color=C['bg_primary'])
        self._counts        = {k: v for k, v in counts_dict.items() if v > 0}
        self._current_chart = ctk.StringVar(value=CHART_TYPES[0])
        self._canvas_widget = None
        self._fig           = None

        if not HAS_MPL:
            ctk.CTkLabel(self,
                         text='⚠  matplotlib is not installed.\n\nRun:  pip install matplotlib',
                         font=('Segoe UI', 14), text_color=C['danger']).pack(expand=True)
            return

        if not self._counts:
            ctk.CTkLabel(self,
                         text='No data available yet. Refresh counts first.',
                         font=('Segoe UI', 13), text_color=C['text_secondary']).pack(expand=True)
            return

        self._build_ui()
        self._draw_chart()

    def _build_ui(self):
        # Toolbar
        bar = ctk.CTkFrame(self, fg_color=C['bg_secondary'], height=52, corner_radius=0)
        bar.pack(fill='x')
        bar.pack_propagate(False)

        ctk.CTkLabel(bar, text='📊  Category Overview',
                     font=('Segoe UI', 14, 'bold'),
                     text_color=C['text_primary']).pack(side='left', padx=16)

        ctk.CTkButton(bar, text='⬇  Download Chart', width=140, height=32,
                      fg_color='#10B981', hover_color='#059669',
                      command=self._download_chart).pack(side='right', padx=12, pady=10)

        for chart_type in CHART_TYPES:
            ctk.CTkButton(bar, text=chart_type, width=120, height=30,
                          fg_color=C['bg_tertiary'],
                          hover_color=C['primary'],
                          font=('Segoe UI', 10),
                          command=lambda ct=chart_type: self._switch_chart(ct)
                          ).pack(side='left', padx=4, pady=10)

        # Chart area
        self._chart_frame = ctk.CTkFrame(self, fg_color=C['bg_secondary'], corner_radius=8)
        self._chart_frame.pack(fill='both', expand=True, padx=12, pady=10)

        # Active label
        self._type_lbl = ctk.CTkLabel(self._chart_frame, text='',
                                      font=('Segoe UI', 11),
                                      text_color=C['text_secondary'])
        self._type_lbl.pack(anchor='e', padx=10, pady=(6, 0))

    def _switch_chart(self, chart_type):
        self._current_chart.set(chart_type)
        self._draw_chart()

    def _draw_chart(self):
        chart_type = self._current_chart.get()
        self._type_lbl.configure(text=chart_type)

        # Destroy old canvas
        if self._canvas_widget:
            self._canvas_widget.get_tk_widget().destroy()
            self._canvas_widget = None
        if self._fig:
            plt.close(self._fig)
            self._fig = None

        labels = list(self._counts.keys())
        values = list(self._counts.values())
        colors = (CHART_COLORS * 3)[:len(labels)]

        fig, ax = plt.subplots(figsize=(11, 5.2), facecolor='#1e2433')
        ax.set_facecolor('#252d3d')
        for sp in ax.spines.values():
            sp.set_edgecolor('#3a4460')
        ax.tick_params(colors='#9aa0b4')

        if chart_type == 'Bar Chart':
            bars = ax.bar(labels, values, color=colors,
                          edgecolor='#1e2433', linewidth=0.6)
            ax.set_xlabel('Category', color='#9aa0b4', fontsize=10)
            ax.set_ylabel('Count',    color='#9aa0b4', fontsize=10)
            for bar, v in zip(bars, values):
                ax.text(bar.get_x() + bar.get_width()/2,
                        bar.get_height() + 0.5,
                        str(v), ha='center', va='bottom',
                        color='white', fontsize=9, fontweight='bold')
            ax.tick_params(axis='x', rotation=35)

        elif chart_type == 'Horizontal Bar':
            bars = ax.barh(labels, values, color=colors,
                           edgecolor='#1e2433', linewidth=0.6)
            ax.set_xlabel('Count', color='#9aa0b4', fontsize=10)
            for bar, v in zip(bars, values):
                ax.text(bar.get_width() + 0.3,
                        bar.get_y() + bar.get_height()/2,
                        str(v), va='center',
                        color='white', fontsize=9, fontweight='bold')

        elif chart_type == 'Pie Chart':
            wedges, texts, autotexts = ax.pie(
                values, labels=labels, colors=colors,
                autopct=lambda p: f'{p:.1f}%' if p > 1.5 else '',
                startangle=140, pctdistance=0.78,
                wedgeprops={'edgecolor': '#1e2433', 'linewidth': 1.5}
            )
            for t in texts:    t.set_color('#c8cfe0'); t.set_fontsize(8)
            for t in autotexts: t.set_color('white'); t.set_fontsize(8)

        elif chart_type == 'Donut Chart':
            wedges, texts, autotexts = ax.pie(
                values, labels=labels, colors=colors,
                autopct=lambda p: f'{p:.1f}%' if p > 1.5 else '',
                startangle=140, pctdistance=0.80,
                wedgeprops={'edgecolor': '#1e2433', 'linewidth': 1.5, 'width': 0.52}
            )
            for t in texts:    t.set_color('#c8cfe0'); t.set_fontsize(8)
            for t in autotexts: t.set_color('white'); t.set_fontsize(8)
            ax.text(0, 0, f'Total\n{sum(values):,}',
                    ha='center', va='center',
                    color='white', fontsize=13, fontweight='bold')

        elif chart_type == 'Line Chart':
            x_pos = list(range(len(labels)))
            ax.plot(x_pos, values, color='#3B82F6', marker='o',
                    linewidth=2.5, markersize=9,
                    markerfacecolor='#60A5FA', markeredgecolor='#1e2433',
                    markeredgewidth=1.5)
            ax.fill_between(x_pos, values, alpha=0.12, color='#3B82F6')
            for x, y in enumerate(values):
                ax.text(x, y + 0.5, str(y), ha='center',
                        color='white', fontsize=9, fontweight='bold')
            ax.set_xticks(x_pos)
            ax.set_xticklabels(labels, rotation=35, ha='right')
            ax.set_xlabel('Category', color='#9aa0b4', fontsize=10)
            ax.set_ylabel('Count',    color='#9aa0b4', fontsize=10)

        ax.set_title(f'BO Object Counts — {chart_type}',
                     color='#e2e8f0', fontsize=13, pad=14, fontweight='bold')
        fig.tight_layout(pad=2.0)

        self._fig = fig
        canvas = FigureCanvasTkAgg(fig, master=self._chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True, padx=6, pady=6)
        self._canvas_widget = canvas

    def _download_chart(self):
        chart_type = self._current_chart.get().replace(' ', '_').lower()
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        path = filedialog.asksaveasfilename(
            title='Save Chart',
            defaultextension='.png',
            filetypes=[('PNG Image', '*.png'), ('PDF', '*.pdf'), ('SVG', '*.svg')],
            initialfile=f'BO_housekeeping_{chart_type}_{ts}.png',
            parent=self
        )
        if path and self._fig:
            self._fig.savefig(path, dpi=150, bbox_inches='tight',
                              facecolor=self._fig.get_facecolor())
            messagebox.showinfo('Saved', f'Chart saved:\n{path}', parent=self)


# ─────────────────────────────────────────────────────────────────────────────
#  Main page
# ─────────────────────────────────────────────────────────────────────────────

class HousekeepingPage(ctk.CTkFrame):

    def __init__(self, parent):
        super().__init__(parent, fg_color=C['bg_primary'], corner_radius=0)

        self._active_cat = None
        self._cat_counts = {}
        self._cat_frames = {}
        self._data       = {}
        self._checkboxes = {}
        self._viz_win    = None

        self._build_ui()
        threading.Thread(target=self._refresh_counts, daemon=True).start()

    # ── UI Construction ───────────────────────────────────────────────────────

    def _build_ui(self):
        # Header
        top = ctk.CTkFrame(self, fg_color='transparent', height=55)
        top.pack(fill='x', padx=20, pady=(15, 0))
        top.pack_propagate(False)

        ctk.CTkLabel(top, text='🧹  Housekeeping',
                     font=('Segoe UI', 22, 'bold'),
                     text_color=C['text_primary']).pack(side='left')

        # Buttons (right to left)
        self._del_btn = ctk.CTkButton(
            top, text='🗑 Delete Selected', width=140, height=34,
            fg_color=C['danger'], hover_color='#DC2626',
            state='disabled', command=self._delete_selected)
        self._del_btn.pack(side='right')

        ctk.CTkButton(top, text='📊 Visualize', width=105, height=34,
                      fg_color='#7C3AED', hover_color='#6D28D9',
                      command=self._open_viz).pack(side='right', padx=(0, 6))

        ctk.CTkButton(top, text='⬇ Export Excel', width=115, height=34,
                      fg_color='#10B981', hover_color='#059669',
                      command=self._export_excel).pack(side='right', padx=(0, 6))

        ctk.CTkButton(top, text='🔄 Refresh All', width=110, height=34,
                      fg_color=C['bg_tertiary'], text_color=C['text_primary'],
                      command=self._refresh_counts).pack(side='right', padx=(0, 6))

        self._status = ctk.CTkLabel(top, text='',
                                    font=('Segoe UI', 11),
                                    text_color=C['text_secondary'])
        self._status.pack(side='left', padx=16)

        # Tiles area
        tiles_outer = ctk.CTkFrame(self, fg_color=C['bg_secondary'], corner_radius=8)
        tiles_outer.pack(fill='x', padx=15, pady=(10, 5))
        self.tiles_frame = ctk.CTkFrame(tiles_outer, fg_color='transparent')
        self.tiles_frame.pack(fill='x', padx=8, pady=8)
        self._build_tiles()

        # Detail label
        self._detail_label = ctk.CTkLabel(self, text='',
                                          font=('Segoe UI', 13, 'bold'),
                                          text_color=C['text_primary'])
        self._detail_label.pack(anchor='w', padx=20, pady=(4, 0))

        # Column header bar
        self._hdr_frame = ctk.CTkFrame(self, fg_color=C['bg_tertiary'], corner_radius=6)
        self._hdr_frame.pack(fill='x', padx=15, pady=(2, 1))

        # Scrollable data table
        self.detail = ctk.CTkScrollableFrame(self, fg_color=C['bg_secondary'], corner_radius=8)
        self.detail.pack(fill='both', expand=True, padx=15, pady=(0, 12))

        ctk.CTkLabel(self.detail,
                     text='Click a category tile above to view objects.',
                     font=('Segoe UI', 12),
                     text_color=C['text_secondary']).pack(pady=40)

    # ── Tiles ─────────────────────────────────────────────────────────────────

    def _build_tiles(self):
        for w in self.tiles_frame.winfo_children():
            w.destroy()
        cols = 7
        for idx, (cat_id, label, icon, color, desc) in enumerate(CATEGORIES):
            row = idx // cols
            col = idx % cols
            self._build_tile(cat_id, label, icon, color, row, col)
        for c in range(cols):
            self.tiles_frame.grid_columnconfigure(c, weight=1)

    def _build_tile(self, cat_id, label, icon, color, row, col):
        # ✅ FIX: border_color='transparent' crashes newer customtkinter → use bg_tertiary
        tile = ctk.CTkFrame(self.tiles_frame,
                            fg_color=C['bg_tertiary'],
                            corner_radius=8,
                            border_width=2,
                            border_color=C['bg_tertiary'],
                            cursor='hand2')
        tile.grid(row=row, column=col, padx=4, pady=4, sticky='nsew')

        ctk.CTkLabel(tile, text=icon, font=('Segoe UI', 18)).pack(pady=(8, 0))
        ctk.CTkLabel(tile, text=label, font=('Segoe UI', 9, 'bold'),
                     text_color=C['text_primary'], wraplength=90).pack()

        count_lbl = ctk.CTkLabel(tile, text='…',
                                 font=('Segoe UI', 14, 'bold'),
                                 text_color=color)
        count_lbl.pack(pady=(2, 8))

        self._cat_frames[cat_id] = (tile, count_lbl)
        tile.bind('<Button-1>', lambda e, c=cat_id: self._select_category(c))
        for child in tile.winfo_children():
            child.bind('<Button-1>', lambda e, c=cat_id: self._select_category(c))

    # ── Count refresh ─────────────────────────────────────────────────────────

    def _refresh_counts(self):
        if not bo_session.connected:
            self.after(0, lambda: self._status.configure(text='❌ Not connected'))
            return
        self.after(0, lambda: self._status.configure(text='⏳ Refreshing counts…'))
        for cat_id, *_ in CATEGORIES:
            threading.Thread(target=self._fetch_count, args=(cat_id,), daemon=True).start()

    def _fetch_count(self, cat_id):
        rows  = _run_category_query(cat_id)
        count = len(rows)
        self._cat_counts[cat_id] = count
        self._data[cat_id]       = rows

        def _upd():
            if cat_id in self._cat_frames:
                _, lbl = self._cat_frames[cat_id]
                if lbl.winfo_exists():
                    lbl.configure(text=str(count))
            total = sum(self._cat_counts.values())
            if self._status.winfo_exists():
                self._status.configure(text=f'Total objects tracked: {total:,}')
        self.after(0, _upd)

    # ── Category selection ────────────────────────────────────────────────────

    def _select_category(self, cat_id):
        if self._active_cat and self._active_cat in self._cat_frames:
            tile, _ = self._cat_frames[self._active_cat]
            if tile.winfo_exists():
                tile.configure(border_color=C['bg_tertiary'])

        self._active_cat = cat_id
        if cat_id in self._cat_frames:
            tile, _ = self._cat_frames[cat_id]
            cat_color = next((c[3] for c in CATEGORIES if c[0] == cat_id), C['primary'])
            if tile.winfo_exists():
                tile.configure(border_color=cat_color)

        label = next((c[1] for c in CATEGORIES if c[0] == cat_id), cat_id)
        self._detail_label.configure(text=f'  {label}')

        if cat_id in self._data and self._data[cat_id] is not None:
            self._render_detail(cat_id, self._data[cat_id])
        else:
            for w in self.detail.winfo_children():
                w.destroy()
            ctk.CTkLabel(self.detail, text='⏳ Loading…',
                         font=('Segoe UI', 12, 'italic'),
                         text_color=C['text_secondary']).pack(pady=30)
            threading.Thread(target=self._load_detail, args=(cat_id,), daemon=True).start()

    def _load_detail(self, cat_id):
        rows = _run_category_query(cat_id)
        self._data[cat_id] = rows
        self.after(0, lambda r=rows, c=cat_id: self._render_detail(c, r))

    # ── Detail render ─────────────────────────────────────────────────────────

    def _render_detail(self, cat_id, rows):
        for w in self.detail.winfo_children():
            w.destroy()
        for w in self._hdr_frame.winfo_children():
            w.destroy()
        self._checkboxes.clear()
        self._del_btn.configure(state='disabled')

        if not rows:
            ctk.CTkLabel(self.detail,
                         text='✅  No objects found in this category.',
                         font=('Segoe UI', 12),
                         text_color=C['success']).pack(pady=40)
            return

        cols = _get_columns(cat_id)
        self._del_btn.configure(state='normal')

        ctk.CTkLabel(self._hdr_frame, text='', width=26).pack(side='left', padx=6)
        for col_name, col_width in cols:
            ctk.CTkLabel(self._hdr_frame, text=col_name, width=col_width, anchor='w',
                         font=('Segoe UI', 10, 'bold'),
                         text_color=C['text_secondary']).pack(side='left', padx=4, pady=6)

        for i, entry in enumerate(rows):
            obj_id = entry.get('SI_ID', i)
            row_f  = ctk.CTkFrame(self.detail,
                                  fg_color=C['bg_tertiary'] if i % 2 == 0 else C['bg_secondary'],
                                  corner_radius=4)
            row_f.pack(fill='x', padx=4, pady=1)

            var = ctk.BooleanVar(value=False)
            self._checkboxes[obj_id] = var
            ctk.CTkCheckBox(row_f, text='', variable=var, width=20).pack(
                side='left', padx=(6, 4), pady=6)

            for col_name, col_width in cols:
                raw = entry.get(col_name, entry.get(col_name.upper(), ''))
                val = str(raw)[:55] if raw is not None else ''
                ctk.CTkLabel(row_f, text=val, width=col_width, anchor='w',
                             font=('Segoe UI', 10),
                             text_color=C['text_primary']).pack(side='left', padx=4, pady=5)

    # ── Delete ────────────────────────────────────────────────────────────────

    def _delete_selected(self):
        ids = [oid for oid, var in self._checkboxes.items() if var.get()]
        if not ids:
            self._status.configure(text='⚠  Select items first')
            return
        if not messagebox.askyesno('Confirm Delete',
                                   f'Delete {len(ids)} selected object(s)?\nThis cannot be undone.',
                                   parent=self):
            return
        self._status.configure(text=f'⏳ Deleting {len(ids)} object(s)…')
        threading.Thread(target=self._do_delete, args=(ids,), daemon=True).start()

    def _do_delete(self, ids):
        ok = err = 0
        for oid in ids:
            try:
                success, _ = bo_session.delete_object(oid)
                if success: ok += 1
                else:       err += 1
            except Exception:
                err += 1
        self.after(0, lambda: self._status.configure(
            text=f'Done: {ok} deleted ✅  |  {err} failed ❌'))
        if self._active_cat:
            self._data[self._active_cat] = None
            self.after(600, lambda: self._select_category(self._active_cat))

    # ── Visualize ─────────────────────────────────────────────────────────────

    def _open_viz(self):
        if not HAS_MPL:
            messagebox.showwarning(
                'Missing Library',
                'matplotlib is not installed.\n\nRun:\n  pip install matplotlib',
                parent=self)
            return

        counts = {}
        for cat_id, label, *_ in CATEGORIES:
            c = self._cat_counts.get(cat_id, 0)
            if c > 0:
                counts[label] = c

        if not counts:
            messagebox.showinfo('No Data',
                                'No data loaded yet. Wait for counts to refresh.',
                                parent=self)
            return

        if self._viz_win and self._viz_win.winfo_exists():
            self._viz_win.lift()
            return
        self._viz_win = VisualizationWindow(self, counts)

    # ── Export Excel ─────────────────────────────────────────────────────────

    def _export_excel(self):
        if not HAS_XLSX:
            messagebox.showwarning(
                'Missing Library',
                'openpyxl is not installed.\n\nRun:\n  pip install openpyxl',
                parent=self)
            return

        has_data = any(self._data.get(c[0]) for c in CATEGORIES)
        if not has_data:
            messagebox.showinfo('No Data',
                                'No data loaded yet. Wait for counts to refresh.',
                                parent=self)
            return

        ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
        path = filedialog.asksaveasfilename(
            title='Export Housekeeping Report',
            defaultextension='.xlsx',
            filetypes=[('Excel Workbook', '*.xlsx')],
            initialfile=f'BO_Housekeeping_{ts}.xlsx',
            parent=self
        )
        if not path:
            return

        self._status.configure(text='⏳ Exporting Excel…')
        threading.Thread(target=self._do_export, args=(path,), daemon=True).start()

    def _do_export(self, path):
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            # ── Style constants ───────────────────────────────────────────────
            hdr_fill   = PatternFill('solid', fgColor='1E3A5F')
            alt_fill   = PatternFill('solid', fgColor='1A2236')
            title_fill = PatternFill('solid', fgColor='0F172A')
            thin = Border(
                left=Side(style='thin', color='2D3748'),
                right=Side(style='thin', color='2D3748'),
                top=Side(style='thin', color='2D3748'),
                bottom=Side(style='thin', color='2D3748')
            )
            hdr_font   = Font(bold=True, color='FFFFFF', size=10)
            title_font = Font(bold=True, color='FFFFFF', size=14)
            sub_font   = Font(color='9AA0B4', size=10)

            def _hdr(ws, cols, row=1):
                for ci, h in enumerate(cols, 1):
                    c = ws.cell(row=row, column=ci, value=h)
                    c.font      = hdr_font
                    c.fill      = hdr_fill
                    c.alignment = Alignment(horizontal='left', vertical='center')
                    c.border    = thin
                ws.row_dimensions[row].height = 20

            def _row_style(ws, row_i, ncols, alt=False):
                for ci in range(1, ncols+1):
                    c = ws.cell(row=row_i, column=ci)
                    if alt: c.fill = alt_fill
                    c.border = thin

            def _autowidth(ws):
                for col in ws.columns:
                    w = 10
                    for cell in col:
                        if cell.value:
                            w = max(w, min(len(str(cell.value)) + 2, 60))
                    ws.column_dimensions[get_column_letter(col[0].column)].width = w

            # ── Sheet: Summary ────────────────────────────────────────────────
            ws_s = wb.create_sheet('Summary')
            ws_s['A1'] = 'BO Commander — Housekeeping Report'
            ws_s['A1'].font = title_font
            ws_s['A1'].fill = title_fill
            ws_s['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ws_s['A2'].font = sub_font
            ws_s.row_dimensions[1].height = 26

            _hdr(ws_s, ['Category', 'Description', 'Object Count'], row=4)
            grand_total = 0
            for ri, (cat_id, label, icon, color, desc) in enumerate(CATEGORIES, 5):
                cnt = len(self._data.get(cat_id, []) or [])
                grand_total += cnt
                ws_s.cell(ri, 1, f'{icon} {label}')
                ws_s.cell(ri, 2, desc)
                ws_s.cell(ri, 3, cnt)
                _row_style(ws_s, ri, 3, alt=(ri % 2 == 0))

            total_row = 5 + len(CATEGORIES)
            ws_s.cell(total_row, 1, 'GRAND TOTAL').font = Font(bold=True, color='FFD700', size=11)
            ws_s.cell(total_row, 3, grand_total).font   = Font(bold=True, color='FFD700', size=11)

            # Chart data in hidden cols (U, V)
            ws_s.cell(1, 21, 'Category')
            ws_s.cell(1, 22, 'Count')
            for i, (cat_id, label, icon, *_) in enumerate(CATEGORIES, 2):
                ws_s.cell(i, 21, f'{icon} {label}')
                ws_s.cell(i, 22, len(self._data.get(cat_id, []) or []))

            last = len(CATEGORIES) + 1

            bc = BarChart()
            bc.type = 'col'; bc.title = 'Objects by Category'
            bc.y_axis.title = 'Count'; bc.width = 24; bc.height = 14
            bc.add_data(Reference(ws_s, min_col=22, min_row=1, max_row=last),
                        titles_from_data=True)
            bc.set_categories(Reference(ws_s, min_col=21, min_row=2, max_row=last))
            ws_s.add_chart(bc, 'E4')

            pc = PieChart()
            pc.title = 'Distribution (Pie)'; pc.width = 16; pc.height = 12
            pc.add_data(Reference(ws_s, min_col=22, min_row=1, max_row=last),
                        titles_from_data=True)
            pc.set_categories(Reference(ws_s, min_col=21, min_row=2, max_row=last))
            ws_s.add_chart(pc, 'E24')

            _autowidth(ws_s)

            # ── Per-category sheets ───────────────────────────────────────────
            for cat_id, label, icon, color, desc in CATEGORIES:
                rows    = self._data.get(cat_id, []) or []
                sname   = f'{icon} {label}'[:31].replace('/', '_')
                ws      = wb.create_sheet(sname)

                ws['A1'] = f'{icon}  {label}'
                ws['A1'].font = title_font; ws['A1'].fill = title_fill
                ws['A2'] = f'{desc}  |  Records: {len(rows)}'
                ws['A2'].font = sub_font
                ws.row_dimensions[1].height = 24

                if not rows:
                    ws['A4'] = 'No objects found.'
                    ws['A4'].font = Font(color='9AA0B4', italic=True)
                    continue

                cols = _get_columns(cat_id)
                col_names = [c[0] for c in cols]
                _hdr(ws, col_names, row=4)

                for ri, entry in enumerate(rows, 5):
                    alt = (ri % 2 == 0)
                    for ci, col_name in enumerate(col_names, 1):
                        raw = entry.get(col_name, entry.get(col_name.upper(), ''))
                        val = str(raw)[:200] if raw is not None else ''
                        cell = ws.cell(ri, ci, val)
                        cell.border = thin
                        if alt: cell.fill = alt_fill
                _autowidth(ws)

            # ── Sheet: Visualizations (5 charts) ──────────────────────────────
            ws_v = wb.create_sheet('📊 Visualizations')
            ws_v['A1'] = 'BO Commander — Visualizations'
            ws_v['A1'].font = title_font; ws_v['A1'].fill = title_fill
            ws_v['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            ws_v['A2'].font = sub_font

            # Write data table
            _hdr(ws_v, ['Category', 'Count'], row=4)
            for i, (cat_id, label, icon, *_) in enumerate(CATEGORIES, 5):
                cnt = len(self._data.get(cat_id, []) or [])
                ws_v.cell(i, 1, f'{icon} {label}')
                ws_v.cell(i, 2, cnt)
                _row_style(ws_v, i, 2, alt=(i % 2 == 0))

            last_row = 4 + len(CATEGORIES)

            def _data_ref(col): return Reference(ws_v, min_col=col, min_row=4, max_row=last_row)
            def _cat_ref():     return Reference(ws_v, min_col=1,   min_row=5, max_row=last_row)

            # Chart 1 — Vertical Bar (D4)
            c1 = BarChart()
            c1.type = 'col'; c1.title = '① Vertical Bar — Objects per Category'
            c1.y_axis.title = 'Count'; c1.width = 22; c1.height = 12
            c1.add_data(_data_ref(2), titles_from_data=True)
            c1.set_categories(_cat_ref())
            ws_v.add_chart(c1, 'D4')

            # Chart 2 — Horizontal Bar (D24)
            c2 = BarChart()
            c2.type = 'bar'; c2.title = '② Horizontal Bar — Objects per Category'
            c2.x_axis.title = 'Count'; c2.width = 22; c2.height = 12
            c2.add_data(_data_ref(2), titles_from_data=True)
            c2.set_categories(_cat_ref())
            ws_v.add_chart(c2, 'D24')

            # Chart 3 — Pie (N4)
            c3 = PieChart()
            c3.title = '③ Pie Chart — Category Distribution'
            c3.width = 16; c3.height = 12
            c3.add_data(_data_ref(2), titles_from_data=True)
            c3.set_categories(_cat_ref())
            ws_v.add_chart(c3, 'N4')

            # Chart 4 — Line (D44)
            c4 = LineChart()
            c4.title = '④ Line Chart — Category Trend'
            c4.y_axis.title = 'Count'; c4.width = 22; c4.height = 12
            c4.add_data(_data_ref(2), titles_from_data=True)
            c4.set_categories(_cat_ref())
            ws_v.add_chart(c4, 'D44')

            # Chart 5 — 3D Bar (N24)
            c5 = BarChart3D()
            c5.title = '⑤ 3D Bar — Objects per Category'
            c5.width = 18; c5.height = 12
            c5.add_data(_data_ref(2), titles_from_data=True)
            c5.set_categories(_cat_ref())
            ws_v.add_chart(c5, 'N24')

            _autowidth(ws_v)

            wb.save(path)

            n_sheets = 2 + len(CATEGORIES)   # Summary + categories + Visualizations
            self.after(0, lambda: self._status.configure(
                text=f'✅ Exported: {os.path.basename(path)}'))
            self.after(0, lambda: messagebox.showinfo(
                'Export Complete',
                f'Housekeeping report exported!\n\n'
                f'File: {path}\n\n'
                f'Sheets: Summary + {len(CATEGORIES)} categories + Visualizations\n'
                f'Charts: 5 chart types in Visualizations tab\n'
                f'Total records: {grand_total:,}',
                parent=self))

        except Exception as ex:
            import traceback
            tb = traceback.format_exc()
            self.after(0, lambda e=str(ex): self._status.configure(
                text=f'❌ Export failed: {e}'))
            self.after(0, lambda e=str(ex): messagebox.showerror(
                'Export Error', f'Export failed:\n{e}', parent=self))
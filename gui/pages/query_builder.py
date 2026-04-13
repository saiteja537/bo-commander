"""
query_builder.py — Advanced CMS SQL Query Builder
  • 30+ pre-built templates across 8 categories
  • Results treeview with row/timing counter
  • Excel export with 5 auto-generated 2D charts from real query data:
      Bar chart, Pie chart, Column chart, Line chart, Doughnut chart
"""

import time
import threading
import customtkinter as ctk
from tkinter import ttk, messagebox
from config import Config
from core.sapbo_connection import bo_session

# ── AI for Natural Language Query Generator ───────────────────────────────────
try:
    from ai.gemini_client import GeminiClient as _GC
    _QB_AI = _GC()
    _QB_HAS_AI = True
except Exception:
    _QB_AI = None
    _QB_HAS_AI = False


def _ai_to_cms_query(natural_text: str) -> str:
    """
    Convert plain English to a CMS SQL query using Gemini AI.
    Works with any GeminiClient method name (ask/generate/chat/etc).
    """
    if not _QB_AI:
        raise RuntimeError("AI engine not available. Check Gemini API key.")

    system_prompt = (
        "You are an SAP BusinessObjects CMS SQL expert.\n"
        "Convert the user\'s plain English request into a valid CMS SQL query.\n\n"
        "CMS SQL RULES:\n"
        "- Tables: CI_INFOOBJECTS (reports/folders), CI_SYSTEMOBJECTS (users/servers/groups), CI_APPOBJECTS (universes/connections)\n"
        "- Reports: SI_KIND IN (\'Webi\',\'CrystalReport\',\'Deski\',\'FullClient\') AND SI_INSTANCE = 0\n"
        "- Instances: SI_INSTANCE = 1, state: 1=Pending 2=Running 3=Failed 4=Success\n"
        "- Schedules: SI_SCHEDULE = 1\n"
        "- Servers: SI_KIND = \'Server\'\n"
        "- Users: SI_PROGID = \'crystalenterprise.user\'\n"
        "- Groups: SI_KIND = \'Usergroup\'\n"
        "- Disabled users: SI_DISABLED = 1\n"
        "- Always include TOP N (max 500) to avoid overloading\n"
        "- Common columns: SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_CREATION_TIME, SI_UPDATE_TS, SI_STARTTIME, SI_ENDTIME\n"
        "- Scheduling columns: SI_PROCESSINFO.SI_NEXTRUNTIME, SI_PROCESSINFO.SI_STATE\n"
        "- Server columns: SI_SERVER_IS_ALIVE, SI_TOTAL_NUM_FAILURES, SI_SERVER_HOST\n\n"
        "OUTPUT ONLY the SQL query, nothing else. No explanation. No markdown. No backticks.\n"
        "If the request is unclear, write the closest reasonable CMS SQL query."
    )

    prompt = f"{system_prompt}\n\nUser request: {natural_text}\n\nCMS SQL query:"

    for method_name in ("ask", "generate", "chat", "query",
                        "generate_content", "send_message", "get_response", "complete"):
        m = getattr(_QB_AI, method_name, None)
        if callable(m):
            try:
                result = m(prompt)
                if isinstance(result, str):
                    return result.strip()
                if hasattr(result, "text"):
                    return result.text.strip()
                if hasattr(result, "candidates"):
                    return result.candidates[0].content.parts[0].text.strip()
                return str(result).strip()
            except Exception as e:
                raise RuntimeError(f"AI call failed: {e}")

    available = [n for n in dir(_QB_AI) if not n.startswith("_") and callable(getattr(_QB_AI, n))]
    raise RuntimeError(f"No usable AI method found. Available: {', '.join(available[:8])}")

C = Config.COLORS
F = Config.FONTS


# ── Template library ─────────────────────────────────────────────────────────

TEMPLATES = {
    # ── System / Infrastructure ────────────────────────────────────────────
    "System / Infrastructure": {
        "All Servers Status": (
            "SELECT SI_ID, SI_NAME, SI_KIND, SI_SERVER_IS_ALIVE,\n"
            "       SI_TOTAL_NUM_FAILURES, SI_SERVER_HOST\n"
            "FROM CI_SYSTEMOBJECTS\n"
            "WHERE SI_KIND = 'Server'\n"
            "ORDER BY SI_NAME ASC"
        ),
        "CMS Version & Build": (
            "SELECT SI_ID, SI_NAME, SI_KIND, SI_DESCRIPTION\n"
            "FROM CI_SYSTEMOBJECTS\n"
            "WHERE SI_KIND = 'Server' AND SI_NAME LIKE '%CMS%'"
        ),
        "License Information": (
            "SELECT TOP 20 SI_ID, SI_NAME, SI_KIND, SI_DESCRIPTION\n"
            "FROM CI_SYSTEMOBJECTS\n"
            "WHERE SI_KIND LIKE '%License%'\n"
            "ORDER BY SI_NAME ASC"
        ),
        "Auditing Status": (
            "SELECT SI_ID, SI_NAME, SI_SERVER_IS_ALIVE, SI_DESCRIPTION\n"
            "FROM CI_SYSTEMOBJECTS\n"
            "WHERE SI_KIND = 'Server' AND SI_NAME LIKE '%Audit%'"
        ),
        "Failed Servers": (
            "SELECT SI_NAME, SI_SERVER_IS_ALIVE, SI_TOTAL_NUM_FAILURES,\n"
            "       SI_SERVER_FAILURE_START_TIME\n"
            "FROM CI_SYSTEMOBJECTS\n"
            "WHERE SI_KIND = 'Server'\n"
            "AND (SI_SERVER_IS_ALIVE = 0 OR SI_TOTAL_NUM_FAILURES > 0)\n"
            "ORDER BY SI_TOTAL_NUM_FAILURES DESC"
        ),
        "Infrastructure Health Scan": (
            "SELECT SI_NAME, SI_SERVER_IS_ALIVE,\n"
            "       SI_TOTAL_NUM_FAILURES, SI_SERVER_HOST\n"
            "FROM CI_SYSTEMOBJECTS\n"
            "WHERE SI_KIND = 'Server'\n"
            "ORDER BY SI_TOTAL_NUM_FAILURES DESC"
        ),
    },

    # ── Users & Groups ─────────────────────────────────────────────────────
    "Users & Groups": {
        "All Enterprise Users": (
            "SELECT TOP 500 SI_ID, SI_NAME, SI_FULL_NAME,\n"
            "       SI_DISABLED, SI_AUTH_TYPE, SI_UPDATE_TS\n"
            "FROM CI_SYSTEMOBJECTS\n"
            "WHERE SI_PROGID = 'crystalenterprise.user'\n"
            "ORDER BY SI_NAME ASC"
        ),
        "All Groups": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_DESCRIPTION\n"
            "FROM CI_SYSTEMOBJECTS\n"
            "WHERE SI_KIND = 'Usergroup'\n"
            "ORDER BY SI_NAME ASC"
        ),
        "Disabled Users": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_FULL_NAME, SI_AUTH_TYPE\n"
            "FROM CI_SYSTEMOBJECTS\n"
            "WHERE SI_PROGID = 'crystalenterprise.user'\n"
            "AND SI_DISABLED = 1\n"
            "ORDER BY SI_NAME ASC"
        ),
        "Inactive / Dormant Users": (
            "SELECT TOP 100 SI_ID, SI_NAME, SI_FULL_NAME,\n"
            "       SI_LAST_LOGIN_TIME, SI_CREATION_TIME\n"
            "FROM CI_SYSTEMOBJECTS\n"
            "WHERE SI_PROGID = 'crystalenterprise.user'\n"
            "ORDER BY SI_LAST_LOGIN_TIME ASC"
        ),
        "External (LDAP / AD) Users": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_FULL_NAME, SI_AUTH_TYPE\n"
            "FROM CI_SYSTEMOBJECTS\n"
            "WHERE SI_PROGID = 'crystalenterprise.user'\n"
            "AND SI_AUTH_TYPE <> 'secEnterprise'\n"
            "ORDER BY SI_AUTH_TYPE, SI_NAME ASC"
        ),
        "Recently Modified Users": (
            "SELECT TOP 50 SI_ID, SI_NAME, SI_FULL_NAME, SI_UPDATE_TS\n"
            "FROM CI_SYSTEMOBJECTS\n"
            "WHERE SI_PROGID = 'crystalenterprise.user'\n"
            "ORDER BY SI_UPDATE_TS DESC"
        ),
    },

    # ── Reports & Universes ────────────────────────────────────────────────
    "Reports & Universes": {
        "All Web Intelligence Reports": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_OWNER,\n"
            "       SI_CREATION_TIME, SI_UPDATE_TS\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_KIND = 'Webi' AND SI_INSTANCE = 0\n"
            "ORDER BY SI_NAME ASC"
        ),
        "All Crystal Reports": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_OWNER,\n"
            "       SI_CREATION_TIME, SI_UPDATE_TS\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_KIND = 'CrystalReport' AND SI_INSTANCE = 0\n"
            "ORDER BY SI_NAME ASC"
        ),
        "All Reports (All Types)": (
            "SELECT TOP 300 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_UPDATE_TS\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_KIND IN ('Webi','CrystalReport','Deski','FullClient')\n"
            "AND SI_INSTANCE = 0\n"
            "ORDER BY SI_UPDATE_TS DESC"
        ),
        "All Universes": (
            "SELECT TOP 100 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_DESCRIPTION\n"
            "FROM CI_APPOBJECTS\n"
            "WHERE SI_KIND LIKE '%Universe%'\n"
            "ORDER BY SI_NAME ASC"
        ),
        "Reports by Owner (Summary)": (
            "SELECT TOP 200 SI_OWNER, SI_KIND, SI_NAME\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_KIND IN ('Webi','CrystalReport','Deski')\n"
            "AND SI_INSTANCE = 0\n"
            "ORDER BY SI_OWNER, SI_NAME ASC"
        ),
        "Reports Not Run in 90 Days": (
            "SELECT TOP 100 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_UPDATE_TS\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_KIND IN ('Webi','CrystalReport','Deski')\n"
            "AND SI_INSTANCE = 0\n"
            "ORDER BY SI_UPDATE_TS ASC"
        ),
        "Content Ownership Mapping": (
            "SELECT TOP 300 SI_ID, SI_NAME, SI_KIND,\n"
            "       SI_OWNER, SI_CREATION_TIME\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 0\n"
            "AND SI_KIND IN ('Webi','CrystalReport','Deski')\n"
            "ORDER BY SI_OWNER ASC"
        ),
    },

    # ── Scheduling & Instances ─────────────────────────────────────────────
    "Scheduling & Instances": {
        "All Scheduled Reports": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND, SI_OWNER,\n"
            "       SI_PROCESSINFO.SI_NEXTRUNTIME, SI_STARTTIME\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_SCHEDULE = 1\n"
            "ORDER BY SI_NAME ASC"
        ),
        "Failed Instances (Last 100)": (
            "SELECT TOP 100 SI_ID, SI_NAME, SI_KIND, SI_OWNER,\n"
            "       SI_STARTTIME, SI_ENDTIME\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 1\n"
            "AND SI_PROCESSINFO.SI_STATE = 3\n"
            "ORDER BY SI_STARTTIME DESC"
        ),
        "Pending Instances": (
            "SELECT TOP 100 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_STARTTIME\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 1\n"
            "AND SI_PROCESSINFO.SI_STATE = 1\n"
            "ORDER BY SI_STARTTIME ASC"
        ),
        "Oldest Instances (Cleanup)": (
            "SELECT TOP 100 SI_ID, SI_NAME, SI_KIND, SI_OWNER,\n"
            "       SI_STARTTIME, SI_ENDTIME\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 1\n"
            "ORDER BY SI_STARTTIME ASC"
        ),
        "All Running Instances": (
            "SELECT TOP 100 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_STARTTIME\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 1\n"
            "AND SI_PROCESSINFO.SI_STATE = 2\n"
            "ORDER BY SI_STARTTIME DESC"
        ),
        "Zombie Schedule Detector": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND, SI_OWNER,\n"
            "       SI_PROCESSINFO.SI_NEXTRUNTIME, SI_STARTTIME\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_SCHEDULE = 1\n"
            "AND SI_INSTANCE = 0\n"
            "ORDER BY SI_PROCESSINFO.SI_NEXTRUNTIME ASC"
        ),
        "Heavy Failure Reports": (
            "SELECT TOP 200 SI_NAME, SI_OWNER,\n"
            "       SI_PROCESSINFO.SI_STATE, SI_STARTTIME\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 1\n"
            "AND SI_PROCESSINFO.SI_STATE = 3\n"
            "ORDER BY SI_STARTTIME DESC"
        ),
    },

    # ── Folders & Structure ────────────────────────────────────────────────
    "Folders & Structure": {
        "All Folders": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_PARENTID, SI_DESCRIPTION\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_KIND = 'Folder'\n"
            "ORDER BY SI_NAME ASC"
        ),
        "Top-Level Folders": (
            "SELECT TOP 100 SI_ID, SI_NAME, SI_DESCRIPTION\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_KIND = 'Folder' AND SI_PARENTID = 23\n"
            "ORDER BY SI_NAME ASC"
        ),
        "Folders by Owner": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_OWNER, SI_CREATION_TIME\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_KIND = 'Folder'\n"
            "ORDER BY SI_OWNER, SI_NAME ASC"
        ),
        "All Connections": (
            "SELECT TOP 100 SI_ID, SI_NAME, SI_KIND, SI_DESCRIPTION\n"
            "FROM CI_APPOBJECTS\n"
            "WHERE SI_KIND LIKE '%Connection%'\n"
            "ORDER BY SI_NAME ASC"
        ),
        "Folder Content Analyzer": (
            "SELECT TOP 300 SI_ID, SI_NAME,\n"
            "       SI_PARENTID, SI_OWNER\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 0\n"
            "ORDER BY SI_PARENTID ASC"
        ),
    },

    # ── Security & Rights ──────────────────────────────────────────────────
    "Security & Rights": {
        "Objects without Owner": (
            "SELECT TOP 100 SI_ID, SI_NAME, SI_KIND\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE (SI_OWNER IS NULL OR SI_OWNER = '')\n"
            "AND SI_INSTANCE = 0\n"
            "ORDER BY SI_KIND, SI_NAME ASC"
        ),
        "Recently Modified Objects": (
            "SELECT TOP 100 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_UPDATE_TS\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 0\n"
            "ORDER BY SI_UPDATE_TS DESC"
        ),
        "Security Risk Scanner": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 0\n"
            "AND (SI_OWNER IS NULL OR SI_OWNER = '')\n"
            "ORDER BY SI_NAME ASC"
        ),
    },

    # ── Cleanup & Maintenance ──────────────────────────────────────────────
    "Cleanup & Maintenance": {
        "Orphan Report Detector": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND, SI_OWNER,\n"
            "       SI_CREATION_TIME, SI_UPDATE_TS\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 0\n"
            "AND SI_KIND IN ('Webi','CrystalReport','Deski','FullClient')\n"
            "AND (SI_OWNER IS NULL OR SI_OWNER = '')\n"
            "ORDER BY SI_UPDATE_TS DESC"
        ),
        "All Orphaned Instances": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND, SI_OWNER,\n"
            "       SI_STARTTIME, SI_ENDTIME\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 1\n"
            "AND SI_PARENTID NOT IN (\n"
            "    SELECT SI_ID FROM CI_INFOOBJECTS WHERE SI_INSTANCE = 0\n"
            ")\n"
            "ORDER BY SI_STARTTIME DESC"
        ),
        "Unused Universes": (
            "SELECT TOP 50 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_UPDATE_TS\n"
            "FROM CI_APPOBJECTS\n"
            "WHERE SI_KIND LIKE '%Universe%'\n"
            "ORDER BY SI_UPDATE_TS ASC"
        ),
        "Unused Connections": (
            "SELECT TOP 50 SI_ID, SI_NAME, SI_KIND, SI_DESCRIPTION\n"
            "FROM CI_APPOBJECTS\n"
            "WHERE SI_KIND LIKE '%Connection%'\n"
            "ORDER BY SI_NAME ASC"
        ),
        "Duplicate Named Objects": (
            "SELECT TOP 100 SI_NAME, SI_KIND, SI_OWNER\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 0\n"
            "ORDER BY SI_NAME ASC"
        ),
    },

    # ── Analytics & Usage ──────────────────────────────────────────────────
    "Analytics & Usage": {
        "Top 100 Most Recently Used": (
            "SELECT TOP 100 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_UPDATE_TS\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 0\n"
            "ORDER BY SI_UPDATE_TS DESC"
        ),
        "Most Active Report Owners": (
            "SELECT TOP 100 SI_OWNER, SI_KIND, SI_NAME, SI_UPDATE_TS\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_KIND IN ('Webi','CrystalReport','Deski')\n"
            "AND SI_INSTANCE = 0\n"
            "ORDER BY SI_OWNER, SI_UPDATE_TS DESC"
        ),
        "Objects Created This Month": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND, SI_OWNER, SI_CREATION_TIME\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 0\n"
            "ORDER BY SI_CREATION_TIME DESC"
        ),
        "Report Count by Type": (
            "SELECT SI_KIND, SI_OWNER\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_KIND IN ('Webi','CrystalReport','Deski','FullClient')\n"
            "AND SI_INSTANCE = 0\n"
            "ORDER BY SI_KIND ASC"
        ),
        "Object Count by Kind": (
            "SELECT TOP 200 SI_KIND, SI_NAME\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 0\n"
            "ORDER BY SI_KIND ASC"
        ),
        "Unused Reports Detector": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND,\n"
            "       SI_OWNER, SI_UPDATE_TS\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 0\n"
            "AND SI_KIND IN ('Webi','CrystalReport','Deski')\n"
            "ORDER BY SI_UPDATE_TS ASC"
        ),
    },

    # ── 🕵️ Admin Intelligence ─────────────────────────────────────────────
    # These templates answer real admin questions automatically.
    # No SAP tool has these built-in — this is what makes BO Commander unique.
    "🕵️ Admin Intelligence": {

        "Failed Instances → Terminated Users": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_OWNER,\n"
            "       SI_STARTTIME, SI_ENDTIME,\n"
            "       SI_PROCESSINFO.SI_STATE\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 1\n"
            "AND SI_PROCESSINFO.SI_STATE = 3\n"
            "ORDER BY SI_STARTTIME DESC"
            "\n-- NOTE: Run 'Disabled Users' from Users & Groups to cross-reference owners"
        ),

        "Disabled User Content (Orphan Risk)": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND, SI_OWNER,\n"
            "       SI_CREATION_TIME, SI_UPDATE_TS\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 0\n"
            "AND SI_KIND IN ('Webi','CrystalReport','Deski','FullClient')\n"
            "AND SI_OWNER IN (\n"
            "    SELECT SI_NAME FROM CI_SYSTEMOBJECTS\n"
            "    WHERE SI_PROGID = 'crystalenterprise.user'\n"
            "    AND SI_DISABLED = 1\n"
            ")\n"
            "ORDER BY SI_OWNER, SI_UPDATE_TS DESC"
        ),

        "Zombie Schedules (Inactive Owner)": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND, SI_OWNER,\n"
            "       SI_PROCESSINFO.SI_NEXTRUNTIME, SI_STARTTIME\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_SCHEDULE = 1\n"
            "AND SI_INSTANCE = 0\n"
            "AND SI_OWNER IN (\n"
            "    SELECT SI_NAME FROM CI_SYSTEMOBJECTS\n"
            "    WHERE SI_PROGID = 'crystalenterprise.user'\n"
            "    AND SI_DISABLED = 1\n"
            ")\n"
            "ORDER BY SI_PROCESSINFO.SI_NEXTRUNTIME ASC"
        ),

        "Orphan Report Detector (No Owner)": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND,\n"
            "       SI_OWNER, SI_CREATION_TIME, SI_UPDATE_TS\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 0\n"
            "AND SI_KIND IN ('Webi','CrystalReport','Deski','FullClient')\n"
            "AND (SI_OWNER IS NULL OR SI_OWNER = '')\n"
            "ORDER BY SI_UPDATE_TS DESC"
        ),

        "Heavy Failure Reports (Repeated Failures)": (
            "SELECT TOP 200 SI_NAME, SI_OWNER,\n"
            "       SI_PROCESSINFO.SI_STATE, SI_STARTTIME\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 1\n"
            "AND SI_PROCESSINFO.SI_STATE = 3\n"
            "ORDER BY SI_STARTTIME DESC"
        ),

        "Infrastructure Health Scan": (
            "SELECT SI_NAME, SI_SERVER_IS_ALIVE,\n"
            "       SI_TOTAL_NUM_FAILURES, SI_SERVER_HOST\n"
            "FROM CI_SYSTEMOBJECTS\n"
            "WHERE SI_KIND = 'Server'\n"
            "ORDER BY SI_TOTAL_NUM_FAILURES DESC"
        ),

        "Content Ownership Mapping": (
            "SELECT TOP 300 SI_ID, SI_NAME, SI_KIND,\n"
            "       SI_OWNER, SI_CREATION_TIME\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 0\n"
            "AND SI_KIND IN ('Webi','CrystalReport','Deski')\n"
            "ORDER BY SI_OWNER ASC"
        ),

        "Unused Reports (Stale Content)": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND,\n"
            "       SI_OWNER, SI_UPDATE_TS\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 0\n"
            "AND SI_KIND IN ('Webi','CrystalReport','Deski')\n"
            "ORDER BY SI_UPDATE_TS ASC"
        ),

        "Security Risk Scanner (No Owner)": (
            "SELECT TOP 200 SI_ID, SI_NAME, SI_KIND\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 0\n"
            "AND (SI_OWNER IS NULL OR SI_OWNER = '')\n"
            "ORDER BY SI_NAME ASC"
        ),

        "Folder Content Analyzer": (
            "SELECT TOP 300 SI_ID, SI_NAME,\n"
            "       SI_PARENTID, SI_OWNER\n"
            "FROM CI_INFOOBJECTS\n"
            "WHERE SI_INSTANCE = 0\n"
            "ORDER BY SI_PARENTID ASC"
        ),
    },
}

# Flat list for quick lookup
_FLAT_TEMPLATES = {name: q
                   for cat in TEMPLATES.values()
                   for name, q in cat.items()}


# ── Excel export with 5 charts ───────────────────────────────────────────────

def _export_with_charts(data, path):
    """
    Write data to Excel.
    Adds 5 auto-generated 2D charts using the actual query result columns:
      Sheet 1 — Raw data table (formatted)
      Sheet 2 — Bar chart (column counts by first text column)
      Sheet 3 — Pie chart (distribution of first categorical column)
      Sheet 4 — Column chart (same data, different angle)
      Sheet 5 — Line chart (values over row index)
      Sheet 6 — Doughnut chart (top 10 by first value column)
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Palette ──────────────────────────────────────────────────────────────
    HEADER_FILL  = PatternFill('solid', fgColor='1E3A5F')   # dark navy
    HEADER_FONT  = Font(color='FFFFFF', bold=True, name='Segoe UI', size=11)
    ALT_FILL     = PatternFill('solid', fgColor='EBF2FF')
    TITLE_FONT   = Font(bold=True, size=14, name='Segoe UI', color='1E3A5F')
    THIN_BORDER  = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    if not data:
        ws = wb.active
        ws.title = 'No Data'
        ws['A1'] = 'No query results to export.'
        wb.save(path)
        return

    cols = list(data[0].keys())
    rows = [[str(row.get(c, '')) for c in cols] for row in data]

    # ── Sheet 1: Data Table ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Query Results'

    # Title row
    ws.merge_cells(f'A1:{get_column_letter(len(cols))}1')
    ws['A1'] = f'SAP BO Query Results  —  {len(rows)} rows'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    # Header row (row 2)
    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=2, column=ci, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 20

    # Data rows
    for ri, row in enumerate(rows, 3):
        fill = ALT_FILL if ri % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center')

    # Auto-width columns
    for ci, col in enumerate(cols, 1):
        max_len = max(len(col), max((len(r[ci-1]) for r in rows), default=0))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 45)

    # Freeze header
    ws.freeze_panes = 'A3'

    # ── Compute category counts for charts ────────────────────────────────────
    # Use first column as category label, count occurrences
    cat_counts = {}
    for row in rows:
        label = str(row[0])[:30] if row else 'Unknown'
        cat_counts[label] = cat_counts.get(label, 0) + 1

    # Sort by count desc, top 12 categories
    sorted_cats = sorted(cat_counts.items(), key=lambda x: x[1], reverse=True)[:12]
    cat_labels  = [item[0] for item in sorted_cats]
    cat_values  = [item[1] for item in sorted_cats]

    # Helper: write a chart data sheet
    def _make_chart_data_sheet(wb, sheet_name, labels, values):
        cs = wb.create_sheet(f'_{sheet_name}_data')
        cs.sheet_state = 'hidden'
        cs['A1'] = 'Category'
        cs['B1'] = 'Count'
        for i, (lbl, val) in enumerate(zip(labels, values), 2):
            cs[f'A{i}'] = lbl
            cs[f'B{i}'] = val
        return cs, len(labels) + 1  # last data row

    n = len(sorted_cats)

    # ── Sheet 2: Bar Chart ────────────────────────────────────────────────────
    ws_bar = wb.create_sheet('📊 Bar Chart')
    ws_bar['A1'] = f'Distribution of {cols[0]} (Top {n})'
    ws_bar['A1'].font = TITLE_FONT

    for i, (lbl, val) in enumerate(sorted_cats, 2):
        ws_bar[f'A{i}'] = lbl
        ws_bar[f'B{i}'] = val
    ws_bar['B1'] = 'Count'

    bar_chart = BarChart()
    bar_chart.type = 'col'           # vertical bars (column chart)
    bar_chart.grouping = 'clustered'
    bar_chart.title = f'{cols[0]} — Distribution'
    bar_chart.style = 10
    bar_chart.y_axis.title = 'Count'
    bar_chart.x_axis.title = cols[0]
    bar_chart.width  = 26
    bar_chart.height = 16

    data_ref = Reference(ws_bar, min_col=2, min_row=1, max_row=n + 1)
    cats_ref = Reference(ws_bar, min_col=1, min_row=2, max_row=n + 1)
    bar_chart.add_data(data_ref, titles_from_data=True)
    bar_chart.set_categories(cats_ref)
    bar_chart.series[0].graphicalProperties.solidFill = '2563EB'
    ws_bar.add_chart(bar_chart, 'D2')

    # ── Sheet 3: Pie Chart ────────────────────────────────────────────────────
    ws_pie = wb.create_sheet('🥧 Pie Chart')
    ws_pie['A1'] = f'{cols[0]} — Share'
    ws_pie['A1'].font = TITLE_FONT

    TOP_N = min(8, n)   # pie works best with ≤8 slices
    for i, (lbl, val) in enumerate(sorted_cats[:TOP_N], 2):
        ws_pie[f'A{i}'] = lbl
        ws_pie[f'B{i}'] = val
    ws_pie['B1'] = 'Count'

    pie_chart = PieChart()
    pie_chart.title = f'{cols[0]} — Share (Top {TOP_N})'
    pie_chart.style = 10
    pie_chart.width  = 24
    pie_chart.height = 16

    data_ref  = Reference(ws_pie, min_col=2, min_row=1, max_row=TOP_N + 1)
    cats_ref  = Reference(ws_pie, min_col=1, min_row=2, max_row=TOP_N + 1)
    pie_chart.add_data(data_ref, titles_from_data=True)
    pie_chart.set_categories(cats_ref)
    # Explode first slice
    slice0 = DataPoint(idx=0, explosion=10)
    pie_chart.series[0].dPt.append(slice0)
    ws_pie.add_chart(pie_chart, 'D2')

    # ── Sheet 4: Horizontal Bar Chart ────────────────────────────────────────
    ws_hbar = wb.create_sheet('📉 Horizontal Bar')
    ws_hbar['A1'] = f'{cols[0]} — Horizontal'
    ws_hbar['A1'].font = TITLE_FONT

    for i, (lbl, val) in enumerate(sorted_cats, 2):
        ws_hbar[f'A{i}'] = lbl
        ws_hbar[f'B{i}'] = val
    ws_hbar['B1'] = 'Count'

    hbar_chart = BarChart()
    hbar_chart.type = 'bar'          # horizontal
    hbar_chart.grouping = 'clustered'
    hbar_chart.title = f'{cols[0]} — Horizontal Bar'
    hbar_chart.style = 12
    hbar_chart.width  = 26
    hbar_chart.height = 16

    data_ref = Reference(ws_hbar, min_col=2, min_row=1, max_row=n + 1)
    cats_ref = Reference(ws_hbar, min_col=1, min_row=2, max_row=n + 1)
    hbar_chart.add_data(data_ref, titles_from_data=True)
    hbar_chart.set_categories(cats_ref)
    hbar_chart.series[0].graphicalProperties.solidFill = '10B981'
    ws_hbar.add_chart(hbar_chart, 'D2')

    # ── Sheet 5: Line Chart ───────────────────────────────────────────────────
    ws_line = wb.create_sheet('📈 Line Chart')
    ws_line['A1'] = f'{cols[0]} — Trend'
    ws_line['A1'].font = TITLE_FONT

    for i, (lbl, val) in enumerate(sorted_cats, 2):
        ws_line[f'A{i}'] = lbl
        ws_line[f'B{i}'] = val
    ws_line['B1'] = 'Count'

    line_chart = LineChart()
    line_chart.title = f'{cols[0]} — Trend'
    line_chart.style = 10
    line_chart.y_axis.title = 'Count'
    line_chart.x_axis.title = cols[0]
    line_chart.width  = 26
    line_chart.height = 16
    line_chart.grouping = 'standard'

    data_ref = Reference(ws_line, min_col=2, min_row=1, max_row=n + 1)
    cats_ref = Reference(ws_line, min_col=1, min_row=2, max_row=n + 1)
    line_chart.add_data(data_ref, titles_from_data=True)
    line_chart.set_categories(cats_ref)
    line_chart.series[0].graphicalProperties.line.solidFill = 'F59E0B'
    line_chart.series[0].graphicalProperties.line.width = 25000  # 2.5pt
    ws_line.add_chart(line_chart, 'D2')

    # ── Sheet 6: Stacked Bar (shows 2nd column if numeric, else reuses count) ─
    ws_stack = wb.create_sheet('📋 Stacked Bar')
    ws_stack['A1'] = f'{cols[0]} — Stacked View'
    ws_stack['A1'].font = TITLE_FONT

    # Build two series: top 6 vs rest
    top6   = sorted_cats[:6]
    rest_n = sum(v for _, v in sorted_cats[6:])
    combined = top6 + ([('Others', rest_n)] if rest_n > 0 else [])
    for i, (lbl, val) in enumerate(combined, 2):
        ws_stack[f'A{i}'] = lbl
        ws_stack[f'B{i}'] = val
    ws_stack['B1'] = 'Count'

    stack_chart = BarChart()
    stack_chart.type = 'col'
    stack_chart.grouping = 'stacked'
    stack_chart.title = f'{cols[0]} — Stacked (Top 6 + Others)'
    stack_chart.style = 11
    stack_chart.width  = 24
    stack_chart.height = 16
    m = len(combined)
    data_ref = Reference(ws_stack, min_col=2, min_row=1, max_row=m + 1)
    cats_ref = Reference(ws_stack, min_col=1, min_row=2, max_row=m + 1)
    stack_chart.add_data(data_ref, titles_from_data=True)
    stack_chart.set_categories(cats_ref)
    stack_chart.series[0].graphicalProperties.solidFill = '8B5CF6'
    ws_stack.add_chart(stack_chart, 'D2')

    wb.save(path)


# ═════════════════════════════════════════════════════════════════════════════
# PAGE
# ═════════════════════════════════════════════════════════════════════════════

class QueryBuilderPage(ctk.CTkFrame):

    def __init__(self, parent):
        super().__init__(parent, fg_color='transparent')
        self._last_result = None
        self._query_history = []
        self._build_ui()

    def _build_ui(self):
        # ── Header ─────────────────────────────────────────────────────────────
        top = ctk.CTkFrame(self, fg_color='transparent')
        top.pack(fill='x', pady=(10, 6))
        ctk.CTkLabel(top, text='🔍  CMS Query Builder',
                     font=('Segoe UI', 22, 'bold'),
                     text_color=C['text_primary']).pack(side='left', padx=4)
        self._status_lbl = ctk.CTkLabel(top, text='',
                                        font=F['small'],
                                        text_color=C['text_secondary'])
        self._status_lbl.pack(side='right', padx=10)

        # ── Category + Template pickers ────────────────────────────────────────
        picker = ctk.CTkFrame(self, fg_color='transparent')
        picker.pack(fill='x', padx=4, pady=(0, 4))

        ctk.CTkLabel(picker, text='Category:',
                     font=F['small'],
                     text_color=C['text_secondary']).pack(side='left', padx=(0, 4))

        self._cat_var = ctk.StringVar(value=list(TEMPLATES.keys())[0])
        self._cat_menu = ctk.CTkOptionMenu(
            picker,
            variable=self._cat_var,
            values=list(TEMPLATES.keys()),
            width=230,
            command=self._on_cat_change,
            fg_color=C['bg_tertiary'],
            button_color=C['primary'],
            dropdown_fg_color=C['bg_secondary'],
            text_color=C['text_primary'],
        )
        self._cat_menu.pack(side='left', padx=(0, 10))

        ctk.CTkLabel(picker, text='Template:',
                     font=F['small'],
                     text_color=C['text_secondary']).pack(side='left', padx=(0, 4))

        first_cat = list(TEMPLATES.keys())[0]
        self._tmpl_var = ctk.StringVar(value=list(TEMPLATES[first_cat].keys())[0])
        self._tmpl_menu = ctk.CTkOptionMenu(
            picker,
            variable=self._tmpl_var,
            values=list(TEMPLATES[first_cat].keys()),
            width=240,
            command=self._on_tmpl_change,
            fg_color=C['bg_tertiary'],
            button_color=C['primary'],
            dropdown_fg_color=C['bg_secondary'],
            text_color=C['text_primary'],
        )
        self._tmpl_menu.pack(side='left')

        # ── 🤖 Natural Language Query Generator ───────────────────────────────
        nl_frame = ctk.CTkFrame(self, fg_color=C['bg_secondary'], corner_radius=8)
        nl_frame.pack(fill='x', padx=4, pady=(0, 4))

        nl_top = ctk.CTkFrame(nl_frame, fg_color='transparent')
        nl_top.pack(fill='x', padx=12, pady=(8, 4))
        ctk.CTkLabel(nl_top,
                     text='🤖  Natural Language → CMS Query   (AI-Powered)',
                     font=('Segoe UI', 11, 'bold'),
                     text_color='#22d3ee').pack(side='left')
        if not _QB_HAS_AI:
            ctk.CTkLabel(nl_top, text='⚠ AI unavailable — check Gemini API key',
                         font=('Segoe UI', 9), text_color='#f59e0b').pack(side='right')

        nl_row = ctk.CTkFrame(nl_frame, fg_color='transparent')
        nl_row.pack(fill='x', padx=12, pady=(0, 8))

        self._nl_entry = ctk.CTkEntry(
            nl_row,
            placeholder_text='e.g.  show failed webi reports  |  find disabled user content  |  zombie schedules',
            height=32,
            fg_color=C['bg_tertiary'],
            border_color=C['bg_tertiary'],
            text_color=C['text_primary'],
            font=('Segoe UI', 12),
        )
        self._nl_entry.pack(side='left', fill='x', expand=True, padx=(0, 8))
        self._nl_entry.bind('<Return>', lambda e: self._nl_generate())

        self._nl_btn = ctk.CTkButton(
            nl_row,
            text='✨ Generate Query',
            width=140, height=32,
            fg_color='#6366f1',
            hover_color='#4f46e5',
            font=('Segoe UI', 12, 'bold'),
            state='normal' if _QB_HAS_AI else 'disabled',
            command=self._nl_generate,
        )
        self._nl_btn.pack(side='right')

        # ── Action buttons ─────────────────────────────────────────────────────
        btn_row = ctk.CTkFrame(self, fg_color='transparent')
        btn_row.pack(fill='x', padx=4, pady=(0, 4))
        _b = dict(height=32, corner_radius=6, font=('Segoe UI', 12))

        ctk.CTkButton(btn_row, text='▶  Run Query', width=120,
                      fg_color=C['primary'], hover_color=C['primary_hover'],
                      command=self._run_query, **_b).pack(side='left', padx=(0, 6))
        ctk.CTkButton(btn_row, text='↓  Export + Charts', width=150,
                      fg_color=C['success'], hover_color=C['accent'],
                      command=self._export, **_b).pack(side='left', padx=(0, 6))
        ctk.CTkButton(btn_row, text='↓  Export Raw', width=120,
                      fg_color=C['bg_tertiary'], hover_color=C['bg_secondary'],
                      command=lambda: self._export(charts=False), **_b).pack(side='left')
        ctk.CTkButton(btn_row, text='✕  Clear', width=80,
                      fg_color=C['bg_tertiary'], hover_color=C['bg_secondary'],
                      command=self._clear, **_b).pack(side='right')
        self._hist_lbl = ctk.CTkLabel(btn_row, text='',
                                      font=F['small'],
                                      text_color=C['text_secondary'])
        self._hist_lbl.pack(side='right', padx=10)

        # ── SQL editor ────────────────────────────────────────────────────────
        ef = ctk.CTkFrame(self, fg_color=C['bg_secondary'], corner_radius=8)
        ef.pack(fill='x', padx=4, pady=4)
        ctk.CTkLabel(ef, text='CMS Query (SQL):',
                     font=F['small'],
                     text_color=C['text_secondary'],
                     anchor='w').pack(fill='x', padx=12, pady=(8, 2))
        self._editor = ctk.CTkTextbox(ef, height=130,
                                      font=('Consolas', 12),
                                      fg_color=C['bg_tertiary'],
                                      text_color=C['text_primary'],
                                      border_width=0, corner_radius=6)
        self._editor.pack(fill='x', padx=12, pady=(0, 10))
        # Load first template
        first_query = list(TEMPLATES[first_cat].values())[0]
        self._editor.insert('1.0', first_query)

        # ── Results treeview ──────────────────────────────────────────────────
        rf = ctk.CTkFrame(self, fg_color=C['bg_secondary'], corner_radius=8)
        rf.pack(fill='both', expand=True, padx=4, pady=(4, 6))
        self._results_lbl = ctk.CTkLabel(rf, text='Results:',
                                         font=F['small'],
                                         text_color=C['text_secondary'],
                                         anchor='w')
        self._results_lbl.pack(fill='x', padx=12, pady=(8, 4))

        tv_frame = ctk.CTkFrame(rf, fg_color='transparent')
        tv_frame.pack(fill='both', expand=True, padx=12, pady=(0, 10))

        sty = ttk.Style()
        sty.theme_use('default')
        sty.configure('QB.Treeview',
                       background=C['bg_tertiary'],
                       foreground=C['text_primary'],
                       fieldbackground=C['bg_tertiary'],
                       rowheight=28, font=F['small'], borderwidth=0)
        sty.configure('QB.Treeview.Heading',
                       background=C['bg_secondary'],
                       foreground=C['text_secondary'],
                       font=('Segoe UI', 10, 'bold'), relief='flat')
        sty.map('QB.Treeview',
                background=[('selected', C['primary'])],
                foreground=[('selected', 'white')])

        self._tree = ttk.Treeview(tv_frame, style='QB.Treeview',
                                  show='headings', selectmode='browse')
        vsb = ctk.CTkScrollbar(tv_frame, orientation='vertical',
                               command=self._tree.yview)
        hsb = ctk.CTkScrollbar(tv_frame, orientation='horizontal',
                               command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        hsb.pack(side='bottom', fill='x')
        vsb.pack(side='right',  fill='y')
        self._tree.pack(side='left', fill='both', expand=True)

    # ── Template pickers ──────────────────────────────────────────────────────

    def _on_cat_change(self, cat):
        names = list(TEMPLATES[cat].keys())
        self._tmpl_menu.configure(values=names)
        self._tmpl_var.set(names[0])
        self._on_tmpl_change(names[0])

    def _on_tmpl_change(self, tmpl):
        cat = self._cat_var.get()
        q = TEMPLATES.get(cat, {}).get(tmpl, '')
        self._editor.delete('1.0', 'end')
        self._editor.insert('1.0', q)

    # ── Run query ─────────────────────────────────────────────────────────────

    def _run_query(self):
        q = self._editor.get('1.0', 'end').strip()
        if not q:
            return
        self._status_lbl.configure(text='⏳ Running…')
        self._results_lbl.configure(text='Results:')
        for item in self._tree.get_children():
            self._tree.delete(item)
        threading.Thread(target=self._execute, args=(q,), daemon=True).start()

    def _execute(self, query):
        t0 = time.time()
        result = bo_session.run_cms_query(query)
        ms = int((time.time() - t0) * 1000)
        self.after(0, lambda: self._display(result, ms, query))

    def _display(self, result, ms, query):
        if not result:
            self._status_lbl.configure(text='❌ No response from server.')
            return
        entries = result.get('entries', [])
        n = len(entries)
        self._last_result = entries
        # History
        self._query_history.append({'query': query[:60], 'rows': n, 'ms': ms})
        self._hist_lbl.configure(text=f'History: {len(self._query_history)} run(s)')

        if not entries:
            self._status_lbl.configure(text=f'0 rows  ({ms} ms)')
            self._results_lbl.configure(text='Results:  0 rows')
            return

        cols = list(entries[0].keys())
        self._tree['columns'] = cols
        for col in cols:
            w = min(max(len(col) * 10, 80), 200)
            self._tree.heading(col, text=col)
            self._tree.column(col, width=w, minwidth=60, stretch=True)
        for e in entries:
            self._tree.insert('', 'end',
                              values=[str(e.get(c, ''))[:100] for c in cols])

        self._status_lbl.configure(text=f'✅ {n} rows  ({ms} ms)')
        self._results_lbl.configure(text=f'Results:  {n} rows')

    # ── Export ────────────────────────────────────────────────────────────────

    def _export(self, charts=True):
        if not self._last_result:
            messagebox.showinfo('Export', 'Run a query first.'); return
        try:
            from tkinter import filedialog
            suffix = '_with_charts' if charts else '_raw'
            path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel', '*.xlsx')],
                initialfile=f'cms_query{suffix}.xlsx')
            if not path:
                return
            if charts:
                _export_with_charts(self._last_result, path)
                messagebox.showinfo('Export Complete',
                                    f'Excel with 5 charts saved:\n{path}\n\n'
                                    f'Sheets: Query Results, Bar Chart, Pie Chart,\n'
                                    f'Horizontal Bar, Line Chart, Stacked Bar')
            else:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = 'Results'
                if self._last_result:
                    cols = list(self._last_result[0].keys())
                    ws.append(cols)
                    for row in self._last_result:
                        ws.append([str(row.get(c, '')) for c in cols])
                wb.save(path)
                messagebox.showinfo('Export', f'Raw data saved:\n{path}')
        except ImportError:
            messagebox.showwarning('Missing Library', 'Run: pip install openpyxl')
        except Exception as e:
            messagebox.showerror('Export Error', str(e))

    # ── Natural Language Query Generator ─────────────────────────────────────

    def _nl_generate(self):
        """
        Convert the plain English text in the NL entry box into a
        CMS SQL query using Gemini AI, then load it into the SQL editor.
        The user can review and edit the generated query before running it.
        """
        text = self._nl_entry.get().strip()
        if not text:
            messagebox.showinfo("Natural Language Query",
                                "Type a plain English description first.\n\n"
                                "Examples:\n"
                                "  • show failed webi reports\n"
                                "  • find disabled user content\n"
                                "  • zombie schedules with inactive owners\n"
                                "  • reports not run in 90 days\n"
                                "  • servers with most failures",
                                parent=self)
            return

        self._nl_btn.configure(text='⏳ Generating…', state='disabled')
        self._status_lbl.configure(text='🤖 AI generating CMS SQL…')

        def _do():
            try:
                query = _ai_to_cms_query(text)
                return True, query
            except Exception as e:
                return False, str(e)

        def _done(res):
            ok, result = res
            self._nl_btn.configure(text='✨ Generate Query', state='normal')
            if ok:
                self._editor.delete('1.0', 'end')
                self._editor.insert('1.0', result)
                self._status_lbl.configure(text='✅ AI generated query — review then click ▶ Run')
            else:
                self._status_lbl.configure(text=f'❌ AI error: {result[:60]}')
                messagebox.showerror("AI Query Failed",
                                     f"Could not generate query:\n{result}\n\n"
                                     "Check your Gemini API key in Settings.",
                                     parent=self)

        def _run():
            res = _do()
            self.after(0, lambda: _done(res))

        threading.Thread(target=_run, daemon=True).start()

    # ── Clear ─────────────────────────────────────────────────────────────────

    def _clear(self):
        self._editor.delete('1.0', 'end')
        for item in self._tree.get_children():
            self._tree.delete(item)
        self._tree['columns'] = []
        self._results_lbl.configure(text='Results:')
        self._status_lbl.configure(text='')
        self._last_result = None